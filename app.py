from flask import Flask, render_template, redirect, url_for, request, flash, send_file, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
from functools import wraps
from flask_wtf import CSRFProtect
from sqlalchemy.engine.url import make_url
import sys
from sqlalchemy import func, or_
from sqlalchemy import case
import re
import uuid
from unidecode import unidecode
from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode
from io import BytesIO
import unicodedata, time, json
import traceback
import io
from sqlalchemy import text
import time


# Inicialización de la aplicación
app = Flask(__name__)

# Usa una SECRET_KEY fija desde entorno en producción para no invalidar sesiones en cada deploy
app.secret_key = os.getenv('SECRET_KEY', 'cambia-esto-en-produccion')
csrf = CSRFProtect(app)

# Contraseña del panel (usa variable de entorno en prod)
app.config["VV_SECRET"] = os.environ.get("VV_SECRET", "oyKSupE7gpOKd1CL")
# Si VV_EXCEL_PATH no existe, el panel usará automáticamente el último .xlsx/.xls en UPLOAD_FOLDER
app.config["VV_EXCEL_PATH"] = os.environ.get("VV_EXCEL_PATH", "rem_pendientes.xlsx")

# -----------------------------
# Configuración de directorios base
# -----------------------------
basedir = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(basedir, 'uploads')
LOCAL_DATA_DIR = os.path.join(basedir, 'data')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOCAL_DATA_DIR, exist_ok=True)

import socket
from urllib.parse import urlencode, parse_qsl

# ===== Helpers vv: leer TODAS las hojas del Excel y normalizar columnas =====
# Caché en memoria (asegurado)
try:
    _VV_CACHE
except NameError:
    _VV_CACHE = {"mtime": None, "rows": []}


def _vv_slug(s: str) -> str:
    if not isinstance(s, str):
        s = "" if pd.isna(s) else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in s.lower().strip() if ch.isalnum() or ch == " ").replace("  "," ").strip()

_VV_COLMAP = {
    "fecha": "fecha",
    "hora_rem": "fecha",  # ← usa Hora_rem como fecha
    "pedido": "pedido", 
    "no_ped": "pedido",   # ← usa No_ped como pedido
    "remisiones": "remisiones",
    "no_rem": "remisiones",  # ← usa No_rem como remisiones
    "factura de anticipo": "factura_anticipo",
    "fac de anticipo": "factura_anticipo", 
    "factura anticipo": "factura_anticipo",
    "factura de remision": "factura_remision",
    "factura remision": "factura_remision",
    "status": "status",
    "status_rem": "status",  # ← usa Status_rem como status
    "estatus": "status",
    "vendedor": "vendedor",
    "nom_age": "vendedor",   # ← usa Nom_age como vendedor
    "cve_age": "vendedor",   # ← código de agente también puede ser vendedor
}

_VV_COLMAP.update({
    "factura de remisión": "factura_remision",
    "columna u de pedido": "pedido",
    "columna aa de pedido": "vendedor",
})


def _vv_normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza columnas asegurando que sean únicas"""
    rename = {}
    used_names = set()
    
    for i, c in enumerate(df.columns):
        original_col = str(c)
        key = _vv_slug(original_col)
        new_name = _VV_COLMAP.get(key, key)
        
        # Hacer el nombre único si ya existe
        base_name = new_name
        counter = 1
        while new_name in used_names:
            new_name = f"{base_name}_{counter}"
            counter += 1
            
        rename[original_col] = new_name
        used_names.add(new_name)
    
    df = df.rename(columns=rename)
    app.logger.info(f"Columnas normalizadas: {list(df.columns)}")
    
    # Asegurar que tenemos todas las columnas necesarias
    required_columns = ["fecha", "pedido", "remisiones", "factura_anticipo", "factura_remision", "status", "vendedor"]
    for col in required_columns:
        if col not in df.columns:
            df[col] = ""
    
    return df[required_columns]


# ========= _vv_date_only SIN WARNING DE PANDAS ===============================
def _vv_date_only(v) -> str:
    """
    Devuelve fecha en YYYY-MM-DD desde múltiples formatos.
    Evita el warning de pandas cuando la cadena ya es ISO (YYYY-MM-DD).
    """
    from datetime import datetime as _dt, date as _date
    import re as _re

    if v is None:
        return ""
    if isinstance(v, _dt):
        return v.date().isoformat()
    if isinstance(v, _date):
        return v.isoformat()

    s = str(v).strip()
    if not s:
        return ""

    # 1) ISO estricto: YYYY-MM-DD -> parse manual sin pandas
    if _re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        try:
            return _dt.strptime(s, "%Y-%m-%d").date().isoformat()
        except Exception:
            return ""

    # 2) dd/mm/yyyy o d/m/yy (México): usar dayfirst=True
    if _re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", s):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        return "" if pd.isna(dt) else dt.date().isoformat()

    # 3) Fallback flexible
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
    return "" if pd.isna(dt) else dt.date().isoformat()
# =============================================================================


# ========= Helper y ruta para "Eliminar TODO" (con CSRF exempt) ==============
def _vv_hard_reset(purge_uploads: bool = False) -> dict:
    import glob, time, gc
    result = {"store_deleted": False, "cache_cleared": False, "uploads_removed": 0}

    store_path = os.path.join(UPLOAD_FOLDER, "_vv_store.json")
    try:
        if os.path.exists(store_path):
            os.remove(store_path)
            result["store_deleted"] = True
    except Exception as e:
        app.logger.warning(f"No se pudo borrar {store_path}: {e}")

    # limpia caché en memoria
    try:
        if isinstance(_VV_CACHE, dict):
            _VV_CACHE["rows"] = []
            _VV_CACHE["mtime"] = None
            result["cache_cleared"] = True
    except Exception as e:
        app.logger.warning(f"No se pudo limpiar _VV_CACHE: {e}")

    def _try_remove(path: str, retries: int = 6, delay: float = 0.5) -> bool:
        for i in range(retries):
            try:
                os.remove(path)
                return True
            except PermissionError as e:
                # archivo bloqueado; intenta liberar recursos propios y reintenta
                gc.collect()
                time.sleep(delay)
            except Exception as e:
                app.logger.warning(f"No se pudo borrar {path}: {e}")
                return False
        app.logger.warning(f"No se pudo borrar {path}: bloqueado por otro proceso.")
        return False

    if purge_uploads:
        try:
            if os.path.isdir(UPLOAD_FOLDER):
                for pat in ["*.xlsx", "*.xls", "*.csv", "*.json"]:
                    for path in glob.glob(os.path.join(UPLOAD_FOLDER, pat)):
                        if os.path.abspath(UPLOAD_FOLDER) in os.path.abspath(path):
                            ok = _try_remove(path)
                            if ok:
                                result["uploads_removed"] += 1
        except Exception as e:
            app.logger.warning(f"Error al purgar uploads: {e}")

    app.logger.info(f"VV RESET => {result}")
    return result



from flask import request, jsonify
@csrf.exempt
@app.post("/vv/delete-all.json")
def vv_delete_all_json():
    """
    Endpoint que consume el botón '🗑️ Eliminar TODO' del frontend.
    Acepta 'purge_uploads' en el form-data para borrar también archivos subidos.
    CSRF exento porque se llama vía fetch sin token.
    """
    purge = bool(request.form.get("purge_uploads"))
    info = _vv_hard_reset(purge_uploads=purge)
    return jsonify({"ok": True, **info})
# =============================================================================



def debug_excel_structure(path):
    """Función para debug - analiza la estructura del Excel"""
    try:
        app.logger.info("=== DEBUG: ANALIZANDO ESTRUCTURA DEL EXCEL ===")
        
        # Leer sin procesar
        df_raw = pd.read_excel(path, header=None, nrows=10)
        app.logger.info(f"Forma del archivo: {df_raw.shape}")
        app.logger.info("Primeras 5 filas crudas:")
        for i in range(min(5, len(df_raw))):
            app.logger.info(f"Fila {i}: {df_raw.iloc[i].tolist()}")
            
        # Verificar tipos de datos por columna
        app.logger.info("Tipos de datos por columna (primeras filas):")
        for col in range(min(10, df_raw.shape[1])):
            sample_vals = []
            for row in range(min(5, len(df_raw))):
                val = df_raw.iloc[row, col]
                if pd.notna(val):
                    sample_vals.append(f"{type(val).__name__}:'{val}'")
            if sample_vals:
                app.logger.info(f"Col {col}: {', '.join(sample_vals)}")
                
    except Exception as e:
        app.logger.error(f"Error en debug: {e}")

def _vv_clean_val(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

def _vv_join_unique_sorted(vals):
    xs = [_vv_clean_val(v) for v in vals if _vv_clean_val(v)]
    return ", ".join(sorted(set(xs))) if xs else ""

def _vv_status_from(remisiones, fac_anticipo, fac_remision):
    # Regla de prioridad: Facturado > Con anticipo > Surtido > Por surtir
    if fac_remision:
        return "Facturado"
    if fac_anticipo:
        return "Con anticipo"
    if remisiones:
        return "Surtido"
    return "Por surtir"

    
def _vv_guess_header_row(df_raw: pd.DataFrame, scan_rows: int = 20) -> int | None:
    """Busca la fila que contiene encabezados (fecha/pedido/...)."""
    must = ["fecha", "pedido"]  # claves mínimas
    nice = ["remision", "remisiones", "factura", "status", "estatus", "vendedor"]

    def norm(s):
        s = "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s)
        s = unidecode(s).lower().strip()
        return s

    n = min(len(df_raw), scan_rows)
    for i in range(n):
        row = [norm(x) for x in df_raw.iloc[i].tolist()]
        row_text = " | ".join([x for x in row if x])
        
        if not row_text:
            continue
            
        app.logger.debug(f"Fila {i}: {row_text}")
            
        has_must = all(any(m in c for c in row) for m in must)
        has_any_nice = any(any(k in c for c in row) for k in nice)
        
        if has_must and has_any_nice:
            app.logger.info(f"✅ Encabezados encontrados en fila {i}: {row_text}")
            return i
            
        # También considerar si encontramos datos de pedido/remisión en filas siguientes
        if i > 0 and any(x.isdigit() and len(x) >= 4 for x in row if x):  # Buscar números de pedido
            app.logger.info(f"✅ Datos encontrados en fila {i}, encabezados probablemente en fila {i-1}")
            return i - 1
    
    app.logger.warning("❌ No se encontró fila de encabezados")
    return None


def _vv_use_header_row(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Si detecta encabezado embebido, lo aplica; si no, devuelve df_raw tal cual."""
    hdr = _vv_guess_header_row(df_raw)
    df = df_raw
    if hdr is not None and hdr < len(df_raw) - 1:
        headers = df_raw.iloc[hdr].astype(str).tolist()
        df = df_raw.iloc[hdr + 1 :].copy()
        df.columns = headers
    return df


def _vv_pick_excel() -> str | None:
    """Elige el Excel a usar:
       1) Si hay archivos en uploads/, usa el MÁS RECIENTE.
       2) Si no hay, usa VV_EXCEL_PATH (si existe).
    """
    # 1) Último archivo subido en /uploads
    try:
        exts = (".xlsx", ".xls")
        files = [
            os.path.join(UPLOAD_FOLDER, f)
            for f in os.listdir(UPLOAD_FOLDER)
            if f.lower().endswith(exts)
        ]
        if files:
            return max(files, key=os.path.getmtime)  # más reciente
    except Exception:
        pass

    # 2) Ruta fija por variable/setting
    env_path = app.config.get("VV_EXCEL_PATH")
    if env_path and os.path.exists(env_path):
        return env_path

    return None

# ===== Subida de Excel para el panel vv =====
ALLOWED_EXTENSIONS = {"xlsx", "xls"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@csrf.exempt
@app.post("/vv/upload")
@login_required
def vv_upload():
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401

    try:
        if "file" not in request.files:
            return jsonify({"error": "No se envió archivo"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Nombre de archivo vacío"}), 400

        if not file or not allowed_file(file.filename):
            return jsonify({"error": "Formato no permitido. Solo .xlsx o .xls"}), 400

        # Guardar archivo
        fname = secure_filename(file.filename)
        path = os.path.join(UPLOAD_FOLDER, fname)
        file.save(path)

        # Limpiar caché COMPLETAMENTE
        _VV_CACHE["mtime"] = None
        _VV_CACHE["rows"] = []

        # Procesar inmediatamente y obtener resultados
        test_rows = _vv_load_rows_from_excel()

        if len(test_rows) == 0:
            return jsonify({
                "ok": False,
                "filename": fname,
                "message": "El archivo se subió pero no se encontraron datos válidos. Revisa los logs.",
                "rows_processed": 0,
                "rows_saved": 0
            })

        rows_saved = 0
        try:
            rows_saved = _vv_upsert_consolidado(test_rows)
            app.logger.info(f"[vv] UPSERT {rows_saved} filas en remisiones_consolidadas")
        except Exception as e:
            app.logger.error(f"[vv] Error guardando consolidado en DB: {e}")

        return jsonify({
            "ok": True,
            "filename": fname,
            "message": f"✅ Archivo procesado correctamente. {len(test_rows)} filas (guardadas: {rows_saved}).",
            "rows_processed": len(test_rows),
            "rows_saved": rows_saved
        })

    except Exception as e:
        app.logger.error(f"Error en upload: {str(e)}")
        return jsonify({"error": f"Error: {str(e)}"}), 500


@app.get("/vv/inspect-file")
@login_required
def vv_inspect_file():
    """INSPECCIÓN HIPER-DETALLADA - CRÍTICA"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    xls_path = _vv_pick_excel()
    if not xls_path or not os.path.exists(xls_path):
        return jsonify({"error": "No hay archivo cargado"})
    
    try:
        # Leer TODO el archivo
        df_raw = pd.read_excel(xls_path, sheet_name=0, header=None, dtype=str)
        
        result = {
            "file": os.path.basename(xls_path),
            "shape": f"{df_raw.shape[0]} filas x {df_raw.shape[1]} columnas",
            "complete_structure": {}
        }
        
        # Mostrar TODAS las filas (hasta 30)
        for i in range(min(30, len(df_raw))):
            row_data = {}
            for j in range(len(df_raw.columns)):
                val = df_raw.iloc[i, j]
                if pd.isna(val) or str(val).strip() == "":
                    row_data[f"col_{j}"] = "VACÍO"
                else:
                    row_data[f"col_{j}"] = str(val)
            result["complete_structure"][f"fila_{i}"] = row_data
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": f"No se pudo leer: {str(e)}"})

@app.get("/vv/debug-file")
@login_required
def vv_debug_file():
    """Diagnóstico completo del archivo actual"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    xls_path = _vv_pick_excel()
    if not xls_path or not os.path.exists(xls_path):
        return jsonify({"error": "No hay archivo cargado"})
    
    try:
        with open(xls_path, 'rb') as f:
            content = f.read()
        
        # Análisis completo
        result = {
            "filename": os.path.basename(xls_path),
            "file_size": len(content),
            "file_extension": os.path.splitext(xls_path)[1].lower(),
            "first_10_bytes_hex": content[:10].hex(),
            "first_200_bytes_ascii": content[:200].decode('ascii', errors='replace'),
            "detected_as_html": b'<html' in content[:1000].lower() or b'<!doctype' in content[:1000].lower(),
            "detected_as_xlsx": content.startswith(b'PK\x03\x04'),
            "detected_as_xls": content.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'),
        }
        
        # Si parece HTML, mostrar más detalles
        if result["detected_as_html"]:
            try:
                html_content = content.decode('utf-8', errors='ignore')
                # Buscar tablas en el HTML
                table_count = html_content.lower().count('<table')
                result["html_table_count"] = table_count
                result["html_preview"] = html_content[:1000]  # Primeros 1000 caracteres
            except Exception as e:
                result["html_error"] = str(e)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": str(e)})

@app.get("/vv/file-info")
@login_required
def vv_file_info():
    """Información detallada del archivo actual"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    xls_path = _vv_pick_excel()
    if not xls_path or not os.path.exists(xls_path):
        return jsonify({"error": "No hay archivo cargado"})
    
    try:
        # Leer el archivo completo como binario para análisis
        with open(xls_path, 'rb') as f:
            content = f.read()
        
        result = {
            "filename": os.path.basename(xls_path),
            "file_size": len(content),
            "file_path": xls_path,
            "first_100_bytes_hex": content[:100].hex(),
            "first_100_bytes_ascii": content[:100].decode('ascii', errors='replace'),
            "file_extension": os.path.splitext(xls_path)[1].lower()
        }
        
        # Detectar tipo de archivo por signature
        signatures = {
            b'PK\x03\x04': 'Excel (.xlsx)',
            b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1': 'Excel antiguo (.xls)',
            b'%PDF': 'PDF',
            b'\xFF\xD8\xFF': 'JPEG',
            b'\x89PNG': 'PNG',
            b'<!DOCTYPE': 'HTML',
            b'<?xml': 'XML'
        }
        
        result["detected_type"] = "Desconocido"
        for sig, file_type in signatures.items():
            if content.startswith(sig):
                result["detected_type"] = file_type
                break
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": str(e)})


def _vv_load_rows_from_excel() -> list:
    # ---------- Helpers mínimos (por si no están arriba) ----------
    try:
        _vv_clean_val
    except NameError:
        def _vv_clean_val(x):
            import pandas as pd
            if pd.isna(x):
                return ""
            s = str(x).strip()
            return "" if s.lower() == "nan" else s

    try:
        _vv_join_unique_sorted
    except NameError:
        def _vv_join_unique_sorted(vals):
            xs = [_vv_clean_val(v) for v in vals if _vv_clean_val(v)]
            return ", ".join(sorted(set(xs))) if xs else ""

    try:
        _vv_status_from
    except NameError:
        def _vv_status_from(remisiones, fac_anticipo, fac_remision):
            # Prioridad: Facturado > Con anticipo > Surtido > Por surtir
            if fac_remision:
                return "Facturado"
            if fac_anticipo:
                return "Con anticipo"
            if remisiones:
                return "Surtido"
            return "Por surtir"

    import re, os, json, traceback, warnings
    import pandas as pd
    from datetime import date, datetime
    from unidecode import unidecode

    # Silenciar warning específico de pandas (por si queda algún parseo suelto)
    warnings.filterwarnings(
        "ignore",
        message=r"Parsing dates in %Y-%m-%d format when dayfirst=True",
        category=UserWarning
    )

    # ---------- Persistencia ligera ----------
    STORE_PATH = os.path.join(UPLOAD_FOLDER, "_vv_store.json")

    def _vv_store_load() -> list:
        try:
            if os.path.exists(STORE_PATH):
                with open(STORE_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception:
            pass
        return []

    def _vv_store_save(rows: list) -> None:
        try:
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            with open(STORE_PATH, "w", encoding="utf-8") as f:
                json.dump(rows, f, ensure_ascii=False)
        except Exception:
            pass

    # ---------- Utils ----------
    def _vv_is_date_like(v: object) -> bool:
        if isinstance(v, (date, datetime)):
            return True
        s = str(v).strip()
        return bool(
            re.match(r"^\d{4}-\d{1,2}-\d{1,2}", s) or
            re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", s)
        )

    # Siempre devolver YYYY-MM-DD sin horas, y sin warnings de dayfirst
    def _vv_date_only(v) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        s = str(v).strip()
        if not s:
            return ""
        try:
            # México suele usar dd/mm/yyyy
            dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
            if pd.isna(dt):
                # último intento sin dayfirst por si vino en mm/dd/yyyy o ISO
                dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
            return "" if pd.isna(dt) else dt.date().isoformat()
        except Exception:
            return ""

    def _vv_text(v: object) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    def _vv_norm_pedido_num(v: object) -> str:
        """Extrae solo dígitos del pedido (p.ej. 'MAT-114876' -> '114876')."""
        s = _vv_text(v).replace(",", "")
        if not s:
            return ""
        runs = re.findall(r"\d+", s)
        if runs:
            longest = max(runs, key=len)
            return longest.lstrip("0") or "0"
        return ""

    def _vv_norm_pedido_num_floatsafe(v: object) -> str:
        """Tolera 114876.0 o 1.14876E5."""
        s = _vv_text(v).replace(",", "")
        if not s:
            return ""
        try:
            if re.match(r"^[+-]?\d+(\.\d+)?([eE][+-]?\d+)?$", s):
                return str(int(float(s)))
        except Exception:
            pass
        return _vv_norm_pedido_num(s)

    def _vv_folio_text(v: object) -> str:
        """Para folios: si parece fecha, devuelve vacío."""
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        if _vv_is_date_like(v):
            return ""
        return str(v).strip()

    def _vv_find_col_by_name(df: pd.DataFrame, candidates: list[str], scan_rows: int = 8):
        """Busca índice de columna por encabezado exacto (normalizado)."""
        cands = {unidecode(x).strip().lower() for x in candidates}
        rows = min(scan_rows, len(df))
        for c in range(df.shape[1]):
            for r in range(rows):
                v = df.iloc[r, c]
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                txt = unidecode(str(v)).strip().lower()
                if txt in cands:
                    return c
        return None

    def _open_engine(path: str):
        """Detecta engine a partir de la cabecera binaria."""
        try:
            with open(path, "rb") as _f:
                head = _f.read(8)
        except Exception:
            return None
        if head.startswith(b'PK\x03\x04'):
            return "openpyxl"   # .xlsx
        if head.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            return "xlrd"       # .xls (antiguo)
        return None

    def _find_upload_by_tokens(*token_groups):
        """Devuelve el primer archivo en uploads/ que matchee los tokens (el más reciente)."""
        try:
            files = os.listdir(UPLOAD_FOLDER)
        except Exception:
            return None
        files_sorted = sorted(
            files, key=lambda f: os.path.getmtime(os.path.join(UPLOAD_FOLDER, f)),
            reverse=True
        )
        for f in files_sorted:
            lf = f.lower()
            if not lf.endswith((".xlsx", ".xls")):
                continue
            for tokens in token_groups:
                if all(tok in lf for tok in tokens):
                    return os.path.join(UPLOAD_FOLDER, f)
        return None

    # Fallback si no existe _vv_pick_excel en tu archivo
    try:
        _vv_pick_excel
    except NameError:
        def _vv_pick_excel():
            return _find_upload_by_tokens(("rem_pend",), ("combinado",), ("master",))

    # ---------- Escáner de FACT_ANT (por encabezados o A/AW) ----------
    def _scan_fact_ant_pairs_from_workbook(xls_path: str, engine: str) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        try:
            with pd.ExcelFile(xls_path, engine=engine) as xls:
                pedido_keys = {"pedi_rem", "pedirem", "no_peda", "no_ped", "no ped", "pedido"}

                preferred = [s for s in xls.sheet_names if any(k in s.lower() for k in ["fact_ant", "fac_ant", "anticipo"])]

                def _autodetect_header_row(df0: pd.DataFrame) -> int | None:
                    for r in range(min(15, len(df0))):
                        row_vals = [unidecode(str(x)).strip().lower() for x in df0.iloc[r].tolist()]
                        if any(v == "no_fac" for v in row_vals) and any(v in pedido_keys for v in row_vals):
                            return r
                    return None

                def _collect_from_sheet(sheet_name: str):
                    nonlocal pairs
                    try:
                        df_raw = pd.read_excel(xls_path, sheet_name=sheet_name, header=None, dtype=str, engine=engine)
                    except Exception:
                        return
                    if df_raw.empty:
                        return
                    hdr = _autodetect_header_row(df_raw)
                    if hdr is not None:
                        df = pd.read_excel(xls_path, sheet_name=sheet_name, header=hdr, dtype=str, engine=engine)
                        cols = [unidecode(str(c)).strip().lower() for c in df.columns]
                        if "no_fac" in cols and any(k in cols for k in pedido_keys):
                            cfac = cols.index("no_fac")
                            for key in ["pedi_rem", "pedirem", "no_peda", "no_ped", "no ped", "pedido"]:
                                if key in cols:
                                    cped = cols.index(key)
                                    break
                            for _, row in df.iterrows():
                                fac = _vv_folio_text(row.iloc[cfac]) if cfac < len(row) else ""
                                pid = _vv_norm_pedido_num_floatsafe(row.iloc[cped]) if cped < len(row) else ""
                                if fac and pid and pid != "0":
                                    pairs.append((fac, pid))
                    else:
                        if df_raw.shape[1] > 48:
                            for i in range(df_raw.shape[0]):
                                fac = _vv_folio_text(df_raw.iloc[i, 0])
                                pid = _vv_norm_pedido_num_floatsafe(df_raw.iloc[i, 48])
                                if fac and pid:
                                    pairs.append((fac, pid))

                for s in preferred:
                    _collect_from_sheet(s)
                    if pairs:
                        return pairs

                for s in xls.sheet_names:
                    if s in preferred:
                        continue
                    try:
                        df_raw = pd.read_excel(xls_path, sheet_name=s, header=None, dtype=str, engine=engine)
                    except Exception:
                        continue
                    if df_raw.empty:
                        continue
                    hdr = _autodetect_header_row(df_raw)
                    if hdr is None:
                        continue
                    try:
                        df = pd.read_excel(xls_path, sheet_name=s, header=hdr, dtype=str, engine=engine)
                    except Exception:
                        continue
                    cols = [unidecode(str(c)).strip().lower() for c in df.columns]
                    if "no_fac" in cols and any(k in cols for k in pedido_keys):
                        cfac = cols.index("no_fac")
                        for key in ["pedi_rem", "pedirem", "no_peda", "no_ped", "no ped", "pedido"]:
                            if key in cols:
                                cped = cols.index(key)
                                break
                        for _, row in df.iterrows():
                            fac = _vv_folio_text(row.iloc[cfac]) if cfac < len(row) else ""
                            pid = _vv_norm_pedido_num_floatsafe(row.iloc[cped]) if cped < len(row) else ""
                            if fac and pid and pid != "0":
                                pairs.append((fac, pid))
        except Exception:
            return pairs
        return pairs

    # ---------- Lector robusto de fact_ant (con/sin encabezados) ----------
    def _fa_pairs_from_df(df: pd.DataFrame) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        try:
            dfh = df.copy()
            dfh.columns = [unidecode(str(c)).strip().lower() for c in dfh.columns]
            cols = list(dfh.columns)
            if "no_fac" in cols and any(k in cols for k in ["pedi_rem","pedirem","no_peda","no_ped","no ped","pedido"]):
                cfac = cols.index("no_fac")
                for key in ["pedi_rem","pedirem","no_peda","no_ped","no ped","pedido"]:
                    if key in cols:
                        cped = cols.index(key)
                        break
                for _, row in dfh.iterrows():
                    fac = _vv_folio_text(row.iloc[cfac]) if cfac < len(row) else ""
                    pid = _vv_norm_pedido_num_floatsafe(row.iloc[cped]) if cped < len(row) else ""
                    if fac and pid and pid != "0":
                        pairs.append((fac, pid))
                return pairs
        except Exception:
            pass

        if df.shape[1] > 48:
            for i in range(df.shape[0]):
                fac = _vv_folio_text(df.iloc[i, 0])
                pid = _vv_norm_pedido_num_floatsafe(df.iloc[i, 48])
                if fac and pid:
                    pairs.append((fac, pid))
        return pairs

    # ---------- Merge incremental + persistencia ----------
    def _vv_merge_consolidated(old_rows: list, new_rows: list) -> list:
        def to_dict(rows):
            d = {}
            for r in rows or []:
                pidnum = _vv_norm_pedido_num(r.get('pedido', ''))
                if not pidnum:
                    continue
                d[pidnum] = {
                    'fecha': _vv_clean_val(r.get('fecha','')),
                    'pedido': _vv_clean_val(r.get('pedido','')),
                    'remisiones': _vv_clean_val(r.get('remisiones','')),
                    'factura_anticipo': _vv_clean_val(r.get('factura_anticipo','')),
                    'factura_remision': _vv_clean_val(r.get('factura_remision','')),
                    'status': _vv_clean_val(r.get('status','')),
                    'vendedor': _vv_clean_val(r.get('vendedor','')),
                }
            return d

        def splitset(csv):
            if not csv: return set()
            return {x.strip() for x in csv.split(",") if x.strip()}

        A = to_dict(old_rows)
        B = to_dict(new_rows)
        keys = set(A.keys()) | set(B.keys())
        out = []
        for k in sorted(keys, key=lambda x: int(x) if x.isdigit() else x):
            a = A.get(k, {})
            b = B.get(k, {})
            fecha     = _vv_date_only(b.get('fecha','') or a.get('fecha',''))
            vendedor  = _vv_clean_val(b.get('vendedor','') or a.get('vendedor',''))

            rems = sorted(splitset(a.get('remisiones','')) | splitset(b.get('remisiones','')))
            fasA = sorted(splitset(a.get('factura_anticipo','')) | splitset(b.get('factura_anticipo','')))
            fasR = sorted(splitset(a.get('factura_remision','')) | splitset(b.get('factura_remision','')))

            pedido_label = _vv_clean_val(b.get('pedido','') or a.get('pedido','') or k)
            preferred_status = _vv_clean_val(b.get('status','') or a.get('status',''))
            derived_status = _vv_status_from(", ".join(rems), ", ".join(fasA), ", ".join(fasR))
            status = preferred_status or derived_status

            out.append({
                'fecha': fecha,
                'pedido': pedido_label,
                'remisiones': ", ".join(rems),
                'factura_anticipo': ", ".join(fasA),
                'factura_remision': ", ".join(fasR),
                'status': status,
                'vendedor': vendedor,
            })
        return out

    # ---------- Localiza fuentes ----------
    path_pedidos  = _find_upload_by_tokens(("pedido",), ("pedidos",), ("rem_ped",))
    path_op2      = _find_upload_by_tokens(("op2",), ("remision",), ("remisiones",))
    path_fact_ant = _find_upload_by_tokens(("fact_ant",), ("fac_ant",), ("anticipo",))
    path_fact2    = _find_upload_by_tokens(("fact2",), ("fact", "2"))
    combined_path = _vv_pick_excel()

    # ---------- Caché ----------
    mtimes = []
    for p in [path_pedidos, path_op2, path_fact_ant, path_fact2, combined_path]:
        if p and os.path.exists(p):
            mtimes.append(os.path.getmtime(p))
    cache_mtime = max(mtimes) if mtimes else None

    if (cache_mtime is not None and
        _VV_CACHE["mtime"] == cache_mtime and
        _VV_CACHE["rows"] and len(_VV_CACHE["rows"]) > 0):
        app.logger.info("📦 Devolviendo datos desde caché")
        return _VV_CACHE["rows"]

    persisted_rows = _vv_store_load()

    try:
        # =====================================================================
        # PRIORIDAD 1: Libro combinado con hoja PEDIDO (base) + sueltos para complementar
        # =====================================================================
        if combined_path and os.path.exists(combined_path):
            eng = _open_engine(combined_path)
            if eng:
                try:
                    with pd.ExcelFile(combined_path, engine=eng) as xls:
                        sheet_ped = None
                        for s in xls.sheet_names:
                            if "pedido" in s.lower():
                                sheet_ped = s
                                break

                    if sheet_ped:
                        # Leer sin encabezados y detectar fila de headers
                        df_ped = pd.read_excel(combined_path, sheet_name=sheet_ped,
                                               header=None, dtype=str, engine=eng)

                        header_row_idx = None
                        for i in range(min(10, len(df_ped))):
                            row_text = " | ".join([str(x) for x in df_ped.iloc[i].fillna('').tolist() if str(x).strip()])
                            if any(k in row_text.lower() for k in ['pedido','fecha','vendedor','no_ped','estatus','status']):
                                header_row_idx = i
                                app.logger.info(f"✅ Encabezados PEDIDOS en fila {i}: {row_text}")
                                break
                        if header_row_idx is None:
                            header_row_idx = 0

                        # POSICIONES: A=0 (fecha), B=1 (pedido), C=2 (serie), AA=26 (vendedor)
                        pedidos_list = []
                        start_row = header_row_idx + 1
                        for idx in range(start_row, len(df_ped)):
                            fecha     = _vv_text(df_ped.iloc[idx, 0] if df_ped.shape[1] > 0  else "")
                            pedido    = _vv_norm_pedido_num_floatsafe(df_ped.iloc[idx, 1] if df_ped.shape[1] > 1  else "")
                            serie_raw = _vv_text(df_ped.iloc[idx, 2] if df_ped.shape[1] > 2  else "")
                            vendedor  = _vv_text(df_ped.iloc[idx, 26] if df_ped.shape[1] > 26 else "")
                            if not pedido:
                                continue
                            serie = re.sub(r"[^A-Z]+", "", unidecode(serie_raw).upper()) or "ISI"
                            pedidos_list.append({
                                'fecha': fecha,
                                'pedido': pedido,
                                'serie': serie,
                                'vendedor': vendedor,
                                'status_excel': "",
                                'remisiones': [],
                                'facturas_anticipo': [],
                                'facturas_remision': []
                            })

                        idx_by_pedido = {p['pedido']: p for p in pedidos_list}

                        # === OP2 dentro del combinado (si existe)
                        with pd.ExcelFile(combined_path, engine=eng) as xls2:
                            sheet_op2 = None
                            for s in xls2.sheet_names:
                                sl = s.lower()
                                if "op" in sl or "remision" in sl:
                                    sheet_op2 = s
                                    break
                        if sheet_op2:
                            df_op2 = pd.read_excel(combined_path, sheet_name=sheet_op2,
                                                   header=None, dtype=str, engine=eng)
                            if not df_op2.empty:
                                rem_col = _vv_find_col_by_name(df_op2, ["no_rem","no rem","remision","remisiones"]) or 0
                                ped_col = _vv_find_col_by_name(df_op2, ["no_ped","no ped","pedido"]) or 12
                                by_pedido = {}
                                for i in range(df_op2.shape[0]):
                                    if df_op2.shape[1] <= max(rem_col, ped_col):
                                        break
                                    rem = _vv_folio_text(df_op2.iloc[i, rem_col])
                                    pid = _vv_norm_pedido_num_floatsafe(df_op2.iloc[i, ped_col])
                                    if not pid or pid == "0" or not rem or rem == "0":
                                        continue
                                    by_pedido.setdefault(pid, set()).add(rem)
                                for pid, rems in by_pedido.items():
                                    tgt = idx_by_pedido.get(pid)
                                    if tgt:
                                        for rfolio in sorted(rems):
                                            if rfolio not in tgt['remisiones']:
                                                tgt['remisiones'].append(rfolio)

                        # === FACT_ANT dentro del combinado (si existe)
                        fa_pairs = _scan_fact_ant_pairs_from_workbook(combined_path, eng)
                        for fac, pid in fa_pairs or []:
                            tgt = idx_by_pedido.get(pid)
                            if tgt and fac not in tgt['facturas_anticipo']:
                                tgt['facturas_anticipo'].append(fac)

                        # === FACT2 dentro del combinado (si existe)
                        with pd.ExcelFile(combined_path, engine=eng) as xls3:
                            sheet_f2 = None
                            for s in xls3.sheet_names:
                                sl = s.lower()
                                if "fact2" in sl or ("fact" in sl and "2" in sl):
                                    sheet_f2 = s
                                    break
                        if sheet_f2:
                            df_f2 = pd.read_excel(combined_path, sheet_name=sheet_f2,
                                                  header=None, dtype=str, engine=eng)
                            if not df_f2.empty and df_f2.shape[1] > 48:
                                for i in range(df_f2.shape[0]):
                                    fac = _vv_folio_text(df_f2.iloc[i, 0])
                                    pid = _vv_norm_pedido_num_floatsafe(df_f2.iloc[i, 48])
                                    if not pid or not fac:
                                        continue
                                    tgt = idx_by_pedido.get(pid)
                                    if tgt and fac not in tgt['facturas_remision']:
                                        tgt['facturas_remision'].append(fac)

                        # === COMPLEMENTOS DESDE ARCHIVOS SUELTOS (para independencia del orden)
                        # OP2 suelto
                        if path_op2 and os.path.exists(path_op2):
                            df_op2_s = pd.read_excel(path_op2, header=None, dtype=str)
                            if not df_op2_s.empty:
                                rem_col = _vv_find_col_by_name(df_op2_s, ["no_rem","no rem","remision","remisiones"]) or 0
                                ped_col = _vv_find_col_by_name(df_op2_s, ["no_ped","no ped","pedido"]) or 12
                                by_pedido = {}
                                for i in range(df_op2_s.shape[0]):
                                    if df_op2_s.shape[1] <= max(rem_col, ped_col):
                                        break
                                    rem = _vv_folio_text(df_op2_s.iloc[i, rem_col])
                                    pid = _vv_norm_pedido_num_floatsafe(df_op2_s.iloc[i, ped_col])
                                    if not pid or pid == "0" or not rem or rem == "0":
                                        continue
                                    by_pedido.setdefault(pid, set()).add(rem)
                                for pid, rems in by_pedido.items():
                                    tgt = idx_by_pedido.get(pid)
                                    if tgt:
                                        for rfolio in sorted(rems):
                                            if rfolio not in tgt['remisiones']:
                                                tgt['remisiones'].append(rfolio)

                        # FACT_ANT suelto (además del del libro)
                        if path_fact_ant and os.path.exists(path_fact_ant):
                            try:
                                try:
                                    df_fa_suelto = pd.read_excel(path_fact_ant, header=0, dtype=str)
                                except Exception:
                                    df_fa_suelto = pd.read_excel(path_fact_ant, header=None, dtype=str)
                                fa_pairs_sueltos = _fa_pairs_from_df(df_fa_suelto)
                                for fac, pid in fa_pairs_sueltos:
                                    tgt = idx_by_pedido.get(pid)
                                    if tgt and fac not in tgt['facturas_anticipo']:
                                        tgt['facturas_anticipo'].append(fac)
                            except Exception as _e:
                                app.logger.warning(f"⚠️ No se pudo leer fact_ant suelto: {_e}")

                        # FACT2 suelto
                        if path_fact2 and os.path.exists(path_fact2):
                            df_f2_s = pd.read_excel(path_fact2, header=None, dtype=str)
                            if not df_f2_s.empty and df_f2_s.shape[1] > 48:
                                for i in range(df_f2_s.shape[0]):
                                    fac = _vv_folio_text(df_f2_s.iloc[i, 0])
                                    pid = _vv_norm_pedido_num_floatsafe(df_f2_s.iloc[i, 48])
                                    if not pid or not fac:
                                        continue
                                    tgt = idx_by_pedido.get(pid)
                                    if tgt and fac not in tgt['facturas_remision']:
                                        tgt['facturas_remision'].append(fac)

                        # Consolidado + merge con lo persistido
                        consolidated_rows = []
                        for p in pedidos_list:
                            rem_join = _vv_join_unique_sorted(p['remisiones'])
                            fa_join  = _vv_join_unique_sorted(p['facturas_anticipo'])
                            fr_join  = _vv_join_unique_sorted(p['facturas_remision'])
                            status   = _vv_clean_val(p.get('status_excel','')) or _vv_status_from(rem_join, fa_join, fr_join)
                            pedido_label = f"{(p.get('serie') or '').strip()}-{_vv_clean_val(p['pedido'])}".strip("-")
                            consolidated_rows.append({
                                'fecha': _vv_date_only(p['fecha']),
                                'pedido': pedido_label,
                                'remisiones': rem_join,
                                'factura_anticipo': fa_join,
                                'factura_remision': fr_join,
                                'status': status,
                                'vendedor': _vv_clean_val(p['vendedor']),
                            })

                        consolidated_rows = sorted(consolidated_rows, key=lambda r: r['pedido'])
                        merged_rows = _vv_merge_consolidated(persisted_rows, consolidated_rows)

                        app.logger.info(f"🎉 CONSOLIDADO (combinado + sueltos + merge): {len(merged_rows)}")

                        _VV_CACHE["mtime"] = cache_mtime
                        _VV_CACHE["rows"] = merged_rows
                        _vv_store_save(merged_rows)
                        return merged_rows

                except Exception as e:
                    app.logger.warning(f"⚠️ Falló ruta 'combinado con PEDIDO': {e}")

        # =====================================================================
        # PRIORIDAD 2: Hoja 'rem_pendientes' como base + sueltos (OP2 / FACT_ANT / FACT2)
        # =====================================================================
        base_rows = None
        if combined_path and os.path.exists(combined_path):
            eng = _open_engine(combined_path)
            if eng:
                try:
                    with pd.ExcelFile(combined_path, engine=eng) as xls:
                        lower_names = [s.lower() for s in xls.sheet_names]
                        if 'rem_pendientes' in lower_names:
                            real_name = [s for s in xls.sheet_names if s.lower() == 'rem_pendientes'][0]
                            df_rp = pd.read_excel(combined_path, sheet_name=real_name, dtype=str, engine=eng)
                            df_rp.columns = [unidecode(str(c)).strip().lower() for c in df_rp.columns]
                            needed = ["fecha","pedido","remisiones","factura de anticipo","factura de remision","status","vendedor"]
                            if all(col in df_rp.columns for col in needed):
                                def juniq(series):
                                    xs = [_vv_clean_val(x) for x in series if _vv_clean_val(x)]
                                    return ", ".join(sorted(set(xs))) if xs else ""
                                def first(series):
                                    for x in series:
                                        v = _vv_clean_val(x)
                                        if v:
                                            return v
                                    return ""
                                out = (df_rp.groupby('pedido', as_index=False).agg({
                                    'fecha': first,
                                    'remisiones': juniq,
                                    'factura de anticipo': juniq,
                                    'factura de remision': juniq,
                                    'status': first,
                                    'vendedor': first
                                }))
                                base_rows = []
                                for _, r in out.iterrows():
                                    base_rows.append({
                                        'fecha': _vv_date_only(r.get('fecha','')),
                                        'pedido': _vv_clean_val(r.get('pedido','')),
                                        'remisiones': _vv_clean_val(r.get('remisiones','')),
                                        'factura_anticipo': _vv_clean_val(r.get('factura de anticipo','')),
                                        'factura_remision': _vv_clean_val(r.get('factura de remision','')),
                                        'status': _vv_clean_val(r.get('status','')),
                                        'vendedor': _vv_clean_val(r.get('vendedor','')),
                                    })

                                # FACT_ANT combinado o suelto
                                fa_pairs = _scan_fact_ant_pairs_from_workbook(combined_path, eng)
                                if (not fa_pairs) and path_fact_ant and os.path.exists(path_fact_ant):
                                    try:
                                        try:
                                            df_fa_suelto = pd.read_excel(path_fact_ant, header=0, dtype=str)
                                        except Exception:
                                            df_fa_suelto = pd.read_excel(path_fact_ant, header=None, dtype=str)
                                        fa_pairs = _fa_pairs_from_df(df_fa_suelto)
                                    except Exception as _e:
                                        app.logger.warning(f"⚠️ No se pudo leer fact_ant suelto: {_e}")

                                if fa_pairs and base_rows:
                                    idx_by_num = {}
                                    for row in base_rows:
                                        num = _vv_norm_pedido_num(row['pedido'])
                                        if num:
                                            idx_by_num.setdefault(num, row)
                                    for fac, pid in fa_pairs:
                                        row = idx_by_num.get(pid)
                                        if row:
                                            existing = [x.strip() for x in row['factura_anticipo'].split(",") if x.strip()] if row['factura_anticipo'] else []
                                            merged = sorted(set(existing + [fac]))
                                            row['factura_anticipo'] = ", ".join(merged)

                                # OP2 suelto también complementa esta base
                                if base_rows and path_op2 and os.path.exists(path_op2):
                                    df_op2 = pd.read_excel(path_op2, header=None, dtype=str)
                                    rem_col = _vv_find_col_by_name(df_op2, ["no_rem","no rem","remision","remisiones"]) or 0
                                    ped_col = _vv_find_col_by_name(df_op2, ["no_ped","no ped","pedido"]) or 12
                                    by_pedido = {}
                                    for i in range(df_op2.shape[0]):
                                        if df_op2.shape[1] <= max(rem_col, ped_col):
                                            break
                                        rem = _vv_folio_text(df_op2.iloc[i, rem_col])
                                        pid = _vv_norm_pedido_num_floatsafe(df_op2.iloc[i, ped_col])
                                        if not pid or pid == "0" or not rem or rem == "0":
                                            continue
                                        by_pedido.setdefault(pid, set()).add(rem)
                                    idx_by_num = { _vv_norm_pedido_num(r['pedido']): r for r in base_rows }
                                    for pid, rems in by_pedido.items():
                                        row = idx_by_num.get(pid)
                                        if row:
                                            existing = [x.strip() for x in row['remisiones'].split(",") if x.strip()] if row['remisiones'] else []
                                            merged = sorted(set(existing).union(rems))
                                            row['remisiones'] = ", ".join(merged)

                                # FACT2 suelto
                                if base_rows and path_fact2 and os.path.exists(path_fact2):
                                    df_f2 = pd.read_excel(path_fact2, header=None, dtype=str)
                                    if not df_f2.empty and df_f2.shape[1] > 48:
                                        idx_by_num = { _vv_norm_pedido_num(r['pedido']): r for r in base_rows }
                                        for i in range(df_f2.shape[0]):
                                            fac = _vv_folio_text(df_f2.iloc[i, 0])
                                            pid = _vv_norm_pedido_num_floatsafe(df_f2.iloc[i, 48])
                                            if not pid or not fac:
                                                continue
                                            row = idx_by_num.get(pid)
                                            if row:
                                                existing = [x.strip() for x in row['factura_remision'].split(",") if x.strip()] if row['factura_remision'] else []
                                                merged = sorted(set(existing + [fac]))
                                                row['factura_remision'] = ", ".join(merged)

                except Exception as e:
                    app.logger.warning(f"⚠️ No se pudo usar 'rem_pendientes': {e}")

        if base_rows is not None:
            consolidated_rows = sorted(base_rows, key=lambda r: r['pedido'])
            merged_rows = _vv_merge_consolidated(persisted_rows, consolidated_rows)

            app.logger.info(f"🎉 CONSOLIDADO (rem_pendientes + sueltos + merge): {len(merged_rows)}")
            _VV_CACHE["mtime"] = cache_mtime
            _VV_CACHE["rows"] = merged_rows
            _vv_store_save(merged_rows)
            return merged_rows

        # =====================================================================
        # PRIORIDAD 3: Multi-fuente suelta (PEDIDOS / OP2 / FACT_ANT / FACT2) en cualquier orden
        # =====================================================================
        sheets_data = {"pedidos": pd.DataFrame(), "op2": pd.DataFrame(), "fact_ant": pd.DataFrame(), "fact2": pd.DataFrame()}

        if path_pedidos and os.path.exists(path_pedidos):
            sheets_data["pedidos"] = pd.read_excel(path_pedidos, header=None, dtype=str)
        if path_op2 and os.path.exists(path_op2):
            sheets_data["op2"] = pd.read_excel(path_op2, header=None, dtype=str)
        if path_fact_ant and os.path.exists(path_fact_ant):
            try:
                _tmpfa = pd.read_excel(path_fact_ant, header=0, dtype=str)
                sheets_data["fact_ant"] = _tmpfa
            except Exception:
                sheets_data["fact_ant"] = pd.read_excel(path_fact_ant, header=None, dtype=str)
        if path_fact2 and os.path.exists(path_fact2):
            sheets_data["fact2"] = pd.read_excel(path_fact2, header=None, dtype=str)

        pedidos_list = []

        def _build_master_from_pedidos(df_pedidos: pd.DataFrame):
            nonlocal pedidos_list
            if df_pedidos.empty:
                return False
            header_row_idx = None
            for i in range(min(10, len(df_pedidos))):
                row_text = " | ".join([str(x) for x in df_pedidos.iloc[i].fillna('').tolist() if str(x).strip()])
                if any(k in row_text.lower() for k in ['pedido','fecha','vendedor','no_ped']):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                header_row_idx = 0

            # Fallback por posiciones (igual que combinado)
            start_data_row = header_row_idx + 1
            for idx in range(start_data_row, len(df_pedidos)):
                fecha     = _vv_text(df_pedidos.iloc[idx, 0] if df_pedidos.shape[1] > 0  else "")
                pedido    = _vv_norm_pedido_num_floatsafe(df_pedidos.iloc[idx, 1] if df_pedidos.shape[1] > 1  else "")
                serie_raw = _vv_text(df_pedidos.iloc[idx, 2] if df_pedidos.shape[1] > 2  else "")
                vendedor  = _vv_text(df_pedidos.iloc[idx, 26] if df_pedidos.shape[1] > 26 else "")
                if pedido:
                    pedidos_list.append({
                        'fecha': fecha,
                        'pedido': pedido,
                        'serie': (re.sub(r"[^A-Z]+", "", unidecode(serie_raw).upper()) or "ISI"),
                        'vendedor': vendedor,
                        'status_excel': "",
                        'remisiones': [],
                        'facturas_anticipo': [],
                        'facturas_remision': []
                    })
            return len(pedidos_list) > 0

        def _build_master_provisional():
            nonlocal pedidos_list
            ids = set()
            df_op2 = sheets_data.get('op2', pd.DataFrame())
            if not df_op2.empty:
                ped_col = _vv_find_col_by_name(df_op2, ["no_ped","no ped","pedido"])
                if ped_col is None: ped_col = 12
                for i in range(df_op2.shape[0]):
                    if df_op2.shape[1] > ped_col:
                        pid = _vv_norm_pedido_num_floatsafe(df_op2.iloc[i, ped_col])
                        if pid:
                            ids.add(pid)
            df_fa = sheets_data.get('fact_ant', pd.DataFrame())
            if not df_fa.empty:
                try:
                    pairs = _fa_pairs_from_df(df_fa)
                    for _, pid in pairs:
                        ids.add(pid)
                except Exception:
                    pass
                if df_fa.shape[1] > 48:
                    for i in range(df_fa.shape[0]):
                        pid = _vv_norm_pedido_num_floatsafe(df_fa.iloc[i, 48])
                        if pid:
                            ids.add(pid)
            df_f2 = sheets_data.get('fact2', pd.DataFrame())
            if not df_f2.empty and df_f2.shape[1] > 48:
                for i in range(df_f2.shape[0]):
                    pid = _vv_norm_pedido_num_floatsafe(df_f2.iloc[i, 48])
                    if pid:
                        ids.add(pid)

            for pid in sorted(ids):
                pedidos_list.append({
                    'fecha': "",
                    'pedido': pid,
                    'serie': "ISI",
                    'vendedor': "",
                    'status_excel': "",
                    'remisiones': [],
                    'facturas_anticipo': [],
                    'facturas_remision': []
                })
            return len(pedidos_list) > 0

        have_master = _build_master_from_pedidos(sheets_data["pedidos"])
        if not have_master:
            if not _build_master_provisional():
                app.logger.error("❌ No hay 'PEDIDOS' ni se pudo construir maestro provisional.")
                _VV_CACHE["rows"] = persisted_rows
                return persisted_rows

        # Cruces (OP2 / FACT_ANT / FACT2)
        idx_by_pedido = {p['pedido']: p for p in pedidos_list}

        if not sheets_data['op2'].empty:
            df_op2 = sheets_data['op2']
            rem_col = _vv_find_col_by_name(df_op2, ["no_rem","no rem","remision","remisiones"]) or 0
            ped_col = _vv_find_col_by_name(df_op2, ["no_ped","no ped","pedido"]) or 12
            by_pedido = {}
            for i in range(df_op2.shape[0]):
                if df_op2.shape[1] <= max(rem_col, ped_col):
                    break
                rem = _vv_folio_text(df_op2.iloc[i, rem_col])
                pid = _vv_norm_pedido_num_floatsafe(df_op2.iloc[i, ped_col])
                if not pid or pid == "0" or not rem or rem == "0":
                    continue
                by_pedido.setdefault(pid, set()).add(rem)
            for pid, rems in by_pedido.items():
                tgt = idx_by_pedido.get(pid)
                if tgt:
                    for rfolio in sorted(rems):
                        if rfolio not in tgt['remisiones']:
                            tgt['remisiones'].append(rfolio)

        fa_pairs = []
        if not sheets_data['fact_ant'].empty:
            try:
                fa_pairs = _fa_pairs_from_df(sheets_data['fact_ant'])
            except Exception:
                fa_pairs = []
        if (not fa_pairs) and combined_path and os.path.exists(combined_path):
            eng = _open_engine(combined_path)
            if eng:
                fa_pairs = _scan_fact_ant_pairs_from_workbook(combined_path, eng)
        if fa_pairs:
            for fac, pid in fa_pairs:
                tgt = idx_by_pedido.get(pid)
                if tgt and fac not in tgt['facturas_anticipo']:
                    tgt['facturas_anticipo'].append(fac)

        if not sheets_data['fact2'].empty and sheets_data['fact2'].shape[1] > 48:
            df_f2 = sheets_data['fact2']
            for i in range(df_f2.shape[0]):
                fac = _vv_folio_text(df_f2.iloc[i, 0])
                pid = _vv_norm_pedido_num_floatsafe(df_f2.iloc[i, 48])
                if not pid or not fac:
                    continue
                tgt = idx_by_pedido.get(pid)
                if tgt and fac not in tgt['facturas_remision']:
                    tgt['facturas_remision'].append(fac)

        consolidated_rows = []
        for p in pedidos_list:
            rem_join = _vv_join_unique_sorted(p['remisiones'])
            fa_join  = _vv_join_unique_sorted(p['facturas_anticipo'])
            fr_join  = _vv_join_unique_sorted(p['facturas_remision'])
            status   = _vv_clean_val(p.get('status_excel','')) or _vv_status_from(rem_join, fa_join, fr_join)
            pedido_label = f"{(p.get('serie') or '').strip()}-{_vv_clean_val(p['pedido'])}".strip("-")
            consolidated_rows.append({
                'fecha': _vv_date_only(p['fecha']),
                'pedido': pedido_label,
                'remisiones': rem_join,
                'factura_anticipo': fa_join,
                'factura_remision': fr_join,
                'status': status,
                'vendedor': _vv_clean_val(p['vendedor']),
            })

        consolidated_rows = sorted(consolidated_rows, key=lambda r: r['pedido'])
        merged_rows = _vv_merge_consolidated(persisted_rows, consolidated_rows)

        _VV_CACHE["mtime"] = cache_mtime
        _VV_CACHE["rows"] = merged_rows
        _vv_store_save(merged_rows)
        return merged_rows

    except Exception as e:
        app.logger.error(f"💥 ERROR en procesamiento: {str(e)}")
        app.logger.error(traceback.format_exc())
        rows = _vv_store_load()
        _VV_CACHE["rows"] = rows
        return rows



@app.get("/vv/sheets-info")
@login_required
def vv_sheets_info():
    """Muestra información de todas las hojas del Excel"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    xls_path = _vv_pick_excel()
    if not xls_path or not os.path.exists(xls_path):
        return jsonify({"error": "No hay archivo cargado"})
    
    try:
        excel_file = pd.ExcelFile(xls_path)
        sheets_info = []
        
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(xls_path, sheet_name=sheet_name, nrows=5, header=None)
                sheets_info.append({
                    "name": sheet_name,
                    "shape": f"{df.shape[0]}x{df.shape[1]}",
                    "first_rows": df.fillna('').astype(str).values.tolist()[:3]
                })
            except Exception as e:
                sheets_info.append({
                    "name": sheet_name,
                    "error": str(e)
                })
        
        return jsonify({
            "file": os.path.basename(xls_path),
            "sheets": sheets_info
        })
        
    except Exception as e:
        return jsonify({"error": str(e)})
    
# ===== Rutas panel oculto "vv" =====
@app.get("/vv")
@login_required
def vv_page():
    # Vista del panel (la protección real de datos está en /vv/data + login del modal)
    return render_template("vv.html")

@app.get("/vv/diagnostic")
@login_required
def vv_diagnostic():
    """Diagnóstico completo del sistema"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        xls_path = _vv_pick_excel()
        result = {
            "session_ok": session.get("vv_ok"),
            "file_path": xls_path,
            "file_exists": os.path.exists(xls_path) if xls_path else False,
            "upload_folder": UPLOAD_FOLDER,
            "files_in_upload": []
        }
        
        # Listar archivos en uploads
        try:
            files = os.listdir(UPLOAD_FOLDER)
            for f in files:
                if f.lower().endswith(('.xlsx', '.xls')):
                    full_path = os.path.join(UPLOAD_FOLDER, f)
                    result["files_in_upload"].append({
                        "name": f,
                        "size": os.path.getsize(full_path),
                        "modified": os.path.getmtime(full_path)
                    })
        except Exception as e:
            result["upload_folder_error"] = str(e)
        
        # Análisis del archivo actual
        if xls_path and os.path.exists(xls_path):
            try:
                with open(xls_path, 'rb') as f:
                    header = f.read(50)
                
                result["file_analysis"] = {
                    "size": os.path.getsize(xls_path),
                    "header_hex": header.hex(),
                    "is_xlsx": header.startswith(b'PK\x03\x04'),
                    "is_xls": header.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'),
                    "is_html": b'<html' in header.lower() or b'<!doctype' in header.lower()
                }
                
                # Intentar lectura básica
                try:
                    test_df = pd.read_excel(xls_path, nrows=5, header=None)
                    result["file_analysis"]["pandas_readable"] = True
                    result["file_analysis"]["shape"] = f"{test_df.shape[0]}x{test_df.shape[1]}"
                except Exception as e:
                    result["file_analysis"]["pandas_readable"] = False
                    result["file_analysis"]["pandas_error"] = str(e)
                    
            except Exception as e:
                result["file_analysis_error"] = str(e)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": str(e)})
    
@app.post("/vv/force-reload")
@login_required
def vv_force_reload():
    """Fuerza el recálculo del archivo a usar"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    # Limpiar caché completamente
    _VV_CACHE["mtime"] = None
    _VV_CACHE["rows"] = []
    
    # Re-seleccionar archivo
    xls_path = _vv_pick_excel()
    
    return jsonify({
        "ok": True,
        "selected_file": xls_path,
        "file_exists": os.path.exists(xls_path) if xls_path else False,
        "message": "Caché limpiada y archivo reseleccionado"
    })

@app.get("/vv/debug")
@login_required
def vv_debug():
    """Endpoint para diagnóstico del archivo actual"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    xls_path = _vv_pick_excel()
    if not xls_path:
        return jsonify({"error": "No hay archivo cargado"})
    
    result = {
        "file_path": xls_path,
        "file_exists": os.path.exists(xls_path),
        "file_size": os.path.getsize(xls_path) if os.path.exists(xls_path) else 0,
        "cache_has_data": len(_VV_CACHE["rows"]) > 0
    }
    
    # Leer header del archivo
    try:
        with open(xls_path, 'rb') as f:
            header = f.read(8)
            result["file_header_hex"] = header.hex()
            result["file_header_ascii"] = header.decode('ascii', errors='replace')
    except Exception as e:
        result["header_error"] = str(e)
    
    return jsonify(result)

@csrf.exempt
@app.post("/vv/clear-cache")
@login_required
def vv_clear_cache():
    """Limpia la caché del Excel"""
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401
    
    _VV_CACHE["mtime"] = None
    _VV_CACHE["rows"] = []
    
    return jsonify({"ok": True, "message": "Caché limpiada"})

@csrf.exempt
@app.post("/vv/login")
@login_required
def vv_login():
    try:
        pwd = (request.form.get("password") or "").strip()
        if pwd == app.config["VV_SECRET"]:
            session["vv_ok"] = True
            session.permanent = True
            return ("", 204)
        return ("Contraseña incorrecta", 401)
    except Exception as e:
        app.logger.error(f"Error en vv_login: {str(e)}")
        return ("Error interno", 500)

# --- Helper de filtros por faltantes (colabora con snake_case) ---
def _vv_apply_missing_filters_snake(rows: list[dict], args) -> list[dict]:
    """
    Filtros opcionales por querystring:
      ?missing_anticipo=1   → pedidos sin 'factura_anticipo'
      ?missing_remision=1   → pedidos sin 'remisiones'
      ?missing_facrem=1     → pedidos sin 'factura_remision'
    Se pueden combinar. Si no viene nada, regresa rows sin cambios.
    """
    def _is_empty(v: str) -> bool:
        if not v:
            return True
        # Considera vacío si sólo hay comas/espacios
        return not any(tok.strip() for tok in str(v).split(",") if tok.strip())

    f_ant   = args.get("missing_anticipo") in ("1", "true", "True", "yes", "on")
    f_rem   = args.get("missing_remision") in ("1", "true", "True", "yes", "on")
    f_facrm = args.get("missing_facrem")   in ("1", "true", "True", "yes", "on")

    if not (f_ant or f_rem or f_facrm):
        return rows

    filtered = []
    for r in rows:
        ant  = r.get("factura_anticipo", "")
        rem  = r.get("remisiones", "")
        facr = r.get("factura_remision", "")

        ok = True
        if f_ant   and not _is_empty(ant):   ok = False
        if f_rem   and not _is_empty(rem):   ok = False
        if f_facrm and not _is_empty(facr):  ok = False

        if ok:
            filtered.append(r)
    return filtered


# --- Ruta /vv/data (única definición) ---
@app.get("/vv/data")
@login_required
def vv_data():
    if not session.get("vv_ok"):
        return jsonify({"error": "No autorizado"}), 401

    try:
        rows_db = []
        source_name = None

        # 1) Intentar DB si está lista
        try:
            if db_ready():  # <- tu función que hace connect SELECT 1
                sql = text("""
                    select
                        pedido_label      as pedido,
                        coalesce(to_char(fecha, 'DD/MM/YYYY'), '') as fecha,
                        remisiones,
                        factura_anticipo,
                        factura_remision,
                        status,
                        vendedor
                    from public.remisiones_consolidadas
                    order by pedido_label asc
                """)
                # IMPORTANTE: usar connect() en lugar de begin()
                with db.engine.connect() as conn:
                    rows_db = [dict(r) for r in conn.execute(sql).mappings().all()]
                source_name = "DB (remisiones_consolidadas)"
            else:
                source_name = "DB no disponible"
        except Exception as e:
            # Si la DB cae a mitad, seguimos en fallback
            app.logger.error(f"[vv] DB inaccesible en /vv/data: {e}")
            source_name = "DB caída"

        # 2) Si DB no trajo datos, caemos a Excel y (si hay DB) persistimos sin romper respuesta
        if not rows_db:
            rows_snake = _vv_load_rows_from_excel()  # [{'fecha','pedido',...}]
            rows_db = rows_snake or []
            # Etiqueta de fuente según el motivo
            source_name = "Excel (fallback)" if rows_snake else (source_name or "Sin datos")

            # Intento de persistir SÓLO si la DB está lista; no bloquea respuesta
            try:
                if rows_snake and db_ready():
                    _vv_upsert_consolidado(rows_snake)
                elif rows_snake:
                    source_name = "Excel (cache-db-fail)"
            except Exception as e:
                app.logger.error(f"[vv] Error persistiendo fallback Excel->DB: {e}")
                source_name = "Excel (cache-db-fail)"

        # 3) Normalizar llaves esperadas por el front
        for r in rows_db:
            r.setdefault("fecha", "")
            r.setdefault("pedido", "")
            r.setdefault("remisiones", "")
            r.setdefault("factura_anticipo", "")
            r.setdefault("factura_remision", "")
            r.setdefault("status", "")
            r.setdefault("vendedor", "")

        # 4) Filtros de faltantes
        rows_filtered = _vv_apply_missing_filters_snake(rows_db, request.args)

        # 5) Mapear al formato de 7 columnas del frontend
        out = []
        for r in rows_filtered:
            out.append({
                "FECHA": r.get("fecha", ""),
                "PEDIDO": r.get("pedido", ""),
                "REMISIONES": r.get("remisiones", ""),
                "FACTURA DE ANTICIPO": r.get("factura_anticipo", ""),
                "FACTURA DE REMISIÓN": r.get("factura_remision", ""),
                "STATUS": r.get("status", ""),
                "VENDEDOR": r.get("vendedor", ""),
            })

        return jsonify({
            "ok": True,
            "data": out,
            "source": source_name,
            "count": len(out)
        })

    except Exception as e:
        app.logger.error(f"[vv] Error en /vv/data: {e}")
        app.logger.error(traceback.format_exc())
        return jsonify({
            "ok": False,
            "data": [],
            "error": str(e),
            "source": None
        }), 500


def _vv_upsert_consolidado(rows: list[dict]) -> int:
    """
    Guarda/actualiza en public.remisiones_consolidadas con UPSERT por (pedido_num).
    Devuelve el número de filas afectadas.
    """
    if not rows:
        return 0

    # Convertimos 'ISI-12345' -> pedido_num='12345' y pedido_label='ISI-12345'
    payload = []
    for r in rows:
        label = (r.get("pedido") or "").strip()
        # extrae solo dígitos del label
        only_digits = "".join(ch for ch in label if ch.isdigit())
        if not only_digits:
            # si no hay dígitos, usamos el label completo como “num” para no perder la fila
            only_digits = label
        payload.append({
            "pedido_num": only_digits,
            "pedido_label": label,
            "fecha": (r.get("fecha") or "")[:10],
            "remisiones": r.get("remisiones") or "",
            "factura_anticipo": r.get("factura_anticipo") or "",
            "factura_remision": r.get("factura_remision") or "",
            "status": r.get("status") or "",
            "vendedor": r.get("vendedor") or "",
        })

    sql = text("""
        insert into public.remisiones_consolidadas
            (pedido_num, pedido_label, fecha, remisiones, factura_anticipo, factura_remision, status, vendedor)
        values
            (:pedido_num, :pedido_label, :fecha, :remisiones, :factura_anticipo, :factura_remision, :status, :vendedor)
        on conflict (pedido_num) do update set
            pedido_label      = excluded.pedido_label,
            fecha             = excluded.fecha,
            remisiones        = excluded.remisiones,
            factura_anticipo  = excluded.factura_anticipo,
            factura_remision  = excluded.factura_remision,
            status            = excluded.status,
            vendedor          = excluded.vendedor,
            updated_at        = now()
    """)

    with app.app_context():
        with db.engine.begin() as conn:
            conn.execute(sql, payload)
            return len(payload)


def _mask_url_safe(u: str) -> str:
    try:
        from sqlalchemy.engine.url import make_url
        uu = make_url(u)
        if uu.password:
            return str(uu._replace(password="***"))
        return str(uu)
    except Exception:
        return u

def _ensure_postgres_uri(uri: str) -> str:
    if not uri:
        return uri

    # Normaliza el esquema base
    fixed = uri.replace('postgres://', 'postgresql://', 1)

    # Fuerza psycopg (driver nuevo) – ideal para Python 3.13 y PgBouncer
    if fixed.startswith('postgresql://'):
        fixed = fixed.replace('postgresql://', 'postgresql+psycopg://', 1)

    # Asegura SSL
    if 'sslmode=' not in fixed:
        fixed += ('&' if '?' in fixed else '?') + 'sslmode=require'

    return fixed

# -----------------------------
# Configuración mejorada de la base de datos
# -----------------------------
def configure_database() -> str:
    """
    Producción (Render):
    - Si hay DATABASE_URL -> usa Postgres (con driver correcto + sslmode=require)
    - Si NO hay DATABASE_URL pero existe PORT (señal de Render) -> ERROR (no usar SQLite)

    Desarrollo local:
    - Usa SQLite persistente en ./data/database.db
    """
    db_url = os.environ.get('DATABASE_URL')

    # 1) Producción con DATABASE_URL
    if db_url:
        print("Configuración: PostgreSQL en producción (Render)")
        uri = _ensure_postgres_uri(db_url)
        print(f"SQLAlchemy URI final => {uri}")
        return uri

    # 2) Render sin DATABASE_URL -> error
    if os.environ.get('PORT'):
        raise RuntimeError(
            "DATABASE_URL no está definida en el entorno de Render. "
            "Ve a Settings → Environment y agrega DATABASE_URL con la cadena de conexión de Postgres."
        )

    # 3) Desarrollo local -> SQLite
    db_path = os.path.join(LOCAL_DATA_DIR, 'database.db')
    print(f"Configuración: SQLite local con persistencia -> {db_path}")
    return f"sqlite:///{db_path}"

# -----------------------------
# Config principal
# -----------------------------
print("=== DEBUG: DATABASE_URL (enmascarada) ===")
print("Variable DATABASE_URL:", _mask_url_safe(os.environ.get('DATABASE_URL', '')))
print("=========================================")

try:
    raw_uri = configure_database()
    app.config['SQLALCHEMY_DATABASE_URI'] = raw_uri
    print(f"✅ Usando DB URI: {_mask_url_safe(app.config['SQLALCHEMY_DATABASE_URI'])}")

except RuntimeError as e:
    print(f"❌ Config DB: {e}")
    raise

app.config.update(
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={
        'pool_pre_ping': True,
        'pool_recycle': 300,    # 5 min: evita conexiones rancias en PgBouncer
        'pool_size': 5,
        'max_overflow': 5,
        'pool_timeout': 20,
        # los connect_args solo aplican con create_engine(); con Flask-SQLAlchemy se pasan igual:
        'connect_args': {
            'sslmode': 'require',
            'connect_timeout': 10,
            'keepalives': 1,
            'keepalives_idle': 30,
            'keepalives_interval': 10,
            'keepalives_count': 3,
        },
    },
    UPLOAD_FOLDER=UPLOAD_FOLDER,
    MAX_CONTENT_LENGTH = 64 * 1024 * 1024,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=2),
    SQLALCHEMY_ECHO=False,
    SQLALCHEMY_RECORD_QUERIES=False,
    PROPAGATE_EXCEPTIONS=True,
)

print("=== DEBUG: CONFIGURACIÓN FINAL (enmascarada) ===")
print("SQLALCHEMY_DATABASE_URI:", _mask_url_safe(app.config.get('SQLALCHEMY_DATABASE_URI', '')))
print("================================================")

# -----------------------------
# Inicialización de extensiones
# -----------------------------
db = SQLAlchemy(app)

# 4) Cerrar sesiones SIEMPRE al final del request/app context
@app.teardown_appcontext
def shutdown_session(exception=None):
    try:
        db.session.remove()
    except Exception:
        # Evita que un error aquí rompa el shutdown
        pass

login_manager = LoginManager(app)
login_manager.login_view = 'login'

# --- Healthcheck DB --- 

def db_ready() -> bool:
    """Verifica si la base de datos está lista SIN martillar a PgBouncer."""
    try:
        with db.engine.connect() as conn:
            # Opción 1 (SQLAlchemy 2.x “crudo”)
            conn.exec_driver_sql("SELECT 1")
            # Opción 2 equivalente:
            # conn.execute(text("SELECT 1"))
        return True
    except Exception:
        app.logger.exception("DB no lista / error de conexión")
        time.sleep(0.5)  # pequeño respiro para no agotar el pool en congestión
        return False

# Modelos de base de datos
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    
    @property
    def password(self):
        raise AttributeError('La contraseña no es un atributo legible')
    
    @password.setter
    def password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def verify_password(self, password):
        return check_password_hash(self.password_hash, password)

class Transferencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.String(20))
    banco = db.Column(db.String(50))
    banco_receptor = db.Column(db.String(50))
    monto = db.Column(db.Float)
    referencia = db.Column(db.String(100), unique=True)
    pedido = db.Column(db.String(100))
    factura = db.Column(db.String(100))
    registrado = db.Column(db.String(50))
    esta_registrado = db.Column(db.Boolean, default=False)
    concepto = db.Column(db.String(200))

class Venta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, index=True)
    concepto = db.Column(db.String(255))
    tipo = db.Column(db.String(100))
    subtipo = db.Column(db.String(100))
    cantidad = db.Column(db.Float)
    usuario = db.Column(db.String(100))
    codigo = db.Column(db.String(50))
    num = db.Column(db.String(50))
    no_fac = db.Column(db.String(50))
    no_nota = db.Column(db.String(50))
    cant = db.Column(db.Float)
    cve_age = db.Column(db.String(20))
    nom_cte = db.Column(db.String(100))
    rfc_cte = db.Column(db.String(30))
    des_mon = db.Column(db.String(20))

    uuid_factura = db.Column(db.String(100))
    uuid_nc = db.Column(db.String(100))
    cliente_1 = db.Column(db.String(200))
    forma_de_pago = db.Column(db.String(50))
    metodo_de_pago = db.Column(db.String(50))
    total_2 = db.Column(db.Float)
    pago_1 = db.Column(db.String(50))

def _to_float(x):
    if x is None:
        return 0.0
    s = str(x).strip()
    if s == '' or s.lower() == 'nan':
        return 0.0
    
    # quita moneda y espacios
    s = s.replace('$', '').replace(' ', '')
    
    # (1,234.56) -> -1234.56
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    
    # normaliza miles/decimales (acepta "1.234,56" y "1,234.56")
    # si hay coma y punto, decide por el último como decimal
    if s.count(',') and s.count('.'):
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        s = s.replace(',', '')
    
    try:
        val = float(s)
        return -val if neg else val
    except:
        return 0.0

def _to_date_any(x, dayfirst=True):
    try:
        # Intenta convertir fecha ignorando cualquier hora que pueda venir
        if pd.isna(x):
            return None
        if isinstance(x, datetime):
            return x.date()
        return pd.to_datetime(x, dayfirst=dayfirst, errors='coerce').date()
    except Exception:
        return None

def _mk_referencia_fallback(banco, fecha, monto, concepto, idx):
    base = f"{banco}|{fecha}|{monto:.2f}|{str(concepto)[:30]}|{idx}"
    # cadena estable y "única" para evitar chocar con unique(referencia)
    return "AUTO-" + uuid.uuid5(uuid.NAMESPACE_URL, base).hex[:18].upper()

from sqlalchemy import text
from sqlalchemy.exc import OperationalError
from flask import jsonify

@app.get("/health/db")
def health_db():
    try:
        # Conexión SIN transacción
        with db.engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return jsonify(status="ok"), 200
    except OperationalError:
        # Limpia pool y reintenta una vez
        try:
            db.engine.dispose()
            with db.engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            return jsonify(status="ok_after_retry"), 200
        except Exception as e2:
            return jsonify(status="fail", error=str(e2)), 500
    except Exception as e:
        return jsonify(status="fail", error=str(e)), 500


def _norm_factura(s: str) -> str:
    """
    Normaliza un número de factura/NC para comparar:
    - a string
    - quita espacios
    - mayúsculas
    - deja solo A-Z, 0-9 y guiones
    """
    s = str(s or "").strip().upper()
    return re.sub(r"[^A-Z0-9\-]", "", s)

def _norm_cols(df):
    """
    Normaliza los nombres de columnas, priorizando el texto entre paréntesis.
    """
    import re
    nuevas = []
    for c in df.columns:
        col = str(c).strip().lower()
        # Primero extraer el texto dentro de paréntesis si existe
        match = re.search(r'\(([^)]+)\)', col)
        if match:
            # Usar el texto dentro de paréntesis como nombre principal
            col = match.group(1).strip().lower()
        else:
            # Si no hay paréntesis, quitar cualquier texto entre paréntesis
            col = re.sub(r"\(.*?\)", "", col)
        
        # quitar acentos
        col = (col.replace("á","a").replace("é","e")
               .replace("í","i").replace("ó","o").replace("ú","u"))
        
        # espacios -> underscore (colapsa múltiples espacios)
        col = re.sub(r"\s+", " ", col).strip().replace(" ", "_")
        nuevas.append(col)
    
    df.columns = nuevas
    
    # Diagnóstico: ahora sí, columnas normalizadas
    try:
        app.logger.info(f"[MP] columnas normalizadas: {list(df.columns)}")
        # Si quieres verlo en UI:
        # flash("MP columnas: " + ", ".join(df.columns), "info")
    except Exception:
        pass
    
    return df

# ----------- PARSERS -----------
def parse_mercado_pago(path):
    """
    Parser ultra-robusto para el formato peculiar de Mercado Pago
    """
    try:
        app.logger.info("=== PARSER ULTRA-ROBUSTO PARA MERCADO PAGO ===")
        
        # 1) LEER TODO EL CONTENIDO
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        app.logger.info(f"Contenido completo ({len(content)} caracteres):")
        app.logger.info(repr(content[:500]))  # Primeros 500 caracteres
        
        # 2) ENCONTRAR Y EXTRAER LOS ENCABEZADOS
        header_match = re.search(r'(Fecha de compra \(date_created\).*?Últimos 4 dígitos \(last_four_digits\))', content, re.DOTALL)
        if not header_match:
            # Buscar patrón alternativo de encabezados
            header_match = re.search(r'(Fecha.*?date_created.*?last_four_digits)', content, re.DOTALL | re.IGNORECASE)
        
        if not header_match:
            flash("No se pudieron identificar los encabezados del reporte", "error")
            return []
        
        headers_full = header_match.group(1)
        app.logger.info(f"Encabezados encontrados: {headers_full}")
        
        # 3) EXTRAER NOMBRES DE COLUMNAS INTERNAS (entre paréntesis)
        internal_columns = re.findall(r'\(([^)]+)\)', headers_full)
        app.logger.info(f"Columnas internas: {internal_columns}")
        
        if not internal_columns:
            flash("No se pudieron extraer las columnas del archivo", "error")
            return []
        
        # 4) ENCONTRAR DONDE COMIENZAN LOS DATOS
        data_start = header_match.end()
        data_content = content[data_start:].strip()
        
        # 5) DIVIDIR EN TRANSACCIONES - Buscar por patrones de fecha
        date_pattern = r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}'
        transactions = []
        current_pos = 0
        
        while True:
            # Buscar la próxima fecha
            date_match = re.search(date_pattern, data_content[current_pos:])
            if not date_match:
                break
                
            start_pos = current_pos + date_match.start()
            
            # Buscar la siguiente fecha (fin de esta transacción)
            next_date_match = re.search(date_pattern, data_content[start_pos + 1:])
            if next_date_match:
                end_pos = start_pos + next_date_match.start()
                transaction_data = data_content[start_pos:end_pos].strip()
            else:
                transaction_data = data_content[start_pos:].strip()
                end_pos = len(data_content)
                
            transactions.append(transaction_data)
            current_pos = end_pos
            
            if len(transactions) > 100:  # Límite de seguridad
                break
        
        app.logger.info(f"Encontradas {len(transactions)} transacciones potenciales")
        
        # 6) PROCESAR CADA TRANSACCIÓN
        rows = []
        for i, transaction in enumerate(transactions[:50]):  # Procesar max 50
            try:
                app.logger.debug(f"Procesando transacción {i}: {transaction[:100]}...")
                
                # Extraer fechas
                dates = re.findall(date_pattern, transaction)
                if not dates:
                    continue
                    
                fecha_str = dates[0]
                fecha = _to_date_any(fecha_str, dayfirst=True)
                if not fecha:
                    continue
                
                # Extraer montos - buscar números con decimales
                amount_pattern = r'-?\d+\.\d{2}'
                amounts = re.findall(amount_pattern, transaction)
                if not amounts:
                    continue
                
                # El último monto positivo suele ser el net_received_amount
                montos_positivos = [float(amt) for amt in amounts if float(amt) > 0]
                if not montos_positivos:
                    continue
                    
                monto = max(montos_positivos)
                
                # Determinar status
                status = 'approved'
                if any(word in transaction.lower() for word in ['rejected', 'rechazado', 'cc_rejected']):
                    status = 'rejected'
                    monto = 0.0
                
                # Extraer referencia (operation_id)
                op_id_match = re.search(r'1\d{11}', transaction)
                referencia = op_id_match.group(0) if op_id_match else f"MP-{i}-{fecha.strftime('%Y%m%d')}"
                
                # Extraer banco/emisor
                bancos = [
                    'BBVA', 'American Express', 'Banamex', 'Master', 'Banco Azteca',
                    'Banorte', 'Visa', 'Amex', 'Mastercard', 'Debvisa'
                ]
                banco = 'Mercado Pago'
                for b in bancos:
                    if b.lower() in transaction.lower():
                        banco = b
                        break
                
                # Extraer concepto
                concepto = "Venta Mercado Pago"
                if 'Venta presencial' in transaction:
                    concepto = "Venta presencial"
                if status == 'rejected':
                    concepto = f"[RECHAZADO] {concepto}"
                
                rows.append({
                    'fecha': fecha.isoformat(),
                    'banco': banco,
                    'banco_receptor': 'Mercado Pago',
                    'monto': monto,
                    'referencia': referencia,
                    'concepto': concepto,
                    'pedido': '',
                    'factura': ''
                })
                
                app.logger.info(f"✓ Transacción {i}: {fecha} - {monto} - {banco}")
                
            except Exception as e:
                app.logger.error(f"Error en transacción {i}: {e}")
                continue
        
        # 7) SI TODO FALLA, USAR MÉTODO DE FALLBACK
        if not rows:
            app.logger.warning("Método principal falló, usando fallback...")
            return parse_mercado_pago_fallback(content)
        
        flash(f"✅ Se procesaron {len(rows)} transacciones de Mercado Pago", "success")
        return rows
        
    except Exception as e:
        app.logger.error(f"Error crítico: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        flash("Error procesando el archivo. Contacte al administrador.", "error")
        return []

def parse_mercado_pago_fallback(content):
    """
    Método de fallback ultra-agresivo para extraer datos
    """
    try:
        app.logger.info("=== MÉTODO FALLBACK ACTIVADO ===")
        rows = []
        
        # Buscar directamente patrones de transacción completos
        transaction_pattern = r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}).*?(-?\d+\.\d{2}).*?(-?\d+\.\d{2}).*?(-?\d+\.\d{2}).*?(\d+\.\d{2})'
        matches = re.findall(transaction_pattern, content, re.DOTALL)
        
        for match in matches:
            try:
                fecha_str, _, _, _, net_received = match
                fecha = _to_date_any(fecha_str, dayfirst=True)
                if fecha:
                    monto = _to_float(net_received)
                    rows.append({
                        'fecha': fecha.isoformat(),
                        'banco': 'Mercado Pago',
                        'banco_receptor': 'Mercado Pago',
                        'monto': max(monto, 0.0),
                        'referencia': f"MP-{fecha.strftime('%Y%m%d%H%M')}",
                        'concepto': 'Venta Mercado Pago',
                        'pedido': '',
                        'factura': ''
                    })
            except:
                continue
        
        # Si aún no hay resultados, buscar solo fechas y montos
        if not rows:
            date_pattern = r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})'
            amount_pattern = r'(\d+\.\d{2})'
            dates = re.findall(date_pattern, content)
            amounts = re.findall(amount_pattern, content)
            
            # Emparejar fechas con montos (asumiendo mismo orden)
            for i, (fecha_str, monto_str) in enumerate(zip(dates, amounts)):
                if i >= min(len(dates), len(amounts)):
                    break
                    
                fecha = _to_date_any(fecha_str, dayfirst=True)
                if fecha:
                    monto = _to_float(monto_str)
                    rows.append({
                        'fecha': fecha.isoformat(),
                        'banco': 'Mercado Pago',
                        'banco_receptor': 'Mercado Pago',
                        'monto': max(monto, 0.0),
                        'referencia': f"MP-{fecha.strftime('%Y%m%d%H%M')}",
                        'concepto': 'Venta Mercado Pago',
                        'pedido': '',
                        'factura': ''
                    })
        
        return rows
        
    except Exception as e:
        app.logger.error(f"Error en fallback: {e}")
        return []

def parse_mercado_pago_para_ventas(path):
    """
    Convierte transacciones de Mercado Pago al formato de Ventas
    """
    try:
        from flask_login import current_user  # Importar aquí para evitar circular imports
        
        # Primero procesamos como transferencia para obtener datos básicos
        transacciones = parse_mercado_pago(path)
        if not transacciones:
            return []
        
        ventas = []
        for i, trans in enumerate(transacciones):
            try:
                # Convertir fecha
                fecha = datetime.fromisoformat(trans['fecha']).date()
                
                # Obtener monto (saltar montos <= 0)
                monto = trans['monto']
                if monto <= 0:
                    continue
                
                # Crear venta con campos CORRECTOS para tu modelo Venta
                venta = {
                    'fecha': fecha,
                    'concepto': f"{trans['concepto']} - {trans['banco']}",
                    'tipo': 'MERCADO_PAGO',
                    'subtipo': 'VENTA_ONLINE',
                    'cantidad': monto,  # Usar 'cantidad' no 'cant'
                    'usuario': current_user.username,  # <-- AQUÍ FALTABA LA COMA
                    # Campos específicos del modelo Venta
                    'codigo': f"MP-{trans['referencia']}",
                    'num': str(i + 1).zfill(4),
                    'no_fac': '',  # No hay factura en MP
                    'no_nota': trans['referencia'],
                    'cve_age': 'MP',  # Clave de agente
                    'nom_cte': 'CLIENTE_MERCADO_PAGO',
                    'rfc_cte': 'XAXX010101000',  # RFC genérico
                    'des_mon': 'MXN'  # Moneda
                }
                ventas.append(venta)  # <-- También corregí el typo "yenta" a "venta"
                
            except Exception as e:
                app.logger.error(f"Error convirtiendo transacción {i}: {e}")
                continue
        
        return ventas
        
    except Exception as e:
        app.logger.error(f"Error general: {e}")
        return []

def _find_header_row_banamex(df_raw, max_scan=80):
    """
    Devuelve (row_idx, use_next_row_as_header)
    - Si encuentra una fila con 'detalle de movimientos' (pocas celdas no vacías), marca use_next_row_as_header=True para tomar la siguiente como encabezados reales.
    - Si encuentra directamente 'fecha/descripción/depositos' en la misma fila, usa esa.
    """
    from unidecode import unidecode
    import re
    
    def norm_cell(v):
        s = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
        s = unidecode(s).strip().lower()
        s = re.sub(r"\s+", " ", s)
        return s
    
    nrows = min(len(df_raw), max_scan)
    for i in range(nrows):
        row_vals = [norm_cell(x) for x in df_raw.iloc[i].tolist()]
        non_empty = [x for x in row_vals if x]
        row_text = " | ".join(non_empty)
        
        # 1) Fila "anuncio" (una sola celda larga)
        if "detalle de movimientos" in row_text and len(non_empty) <= 3:
            return i, True
        
        # 2) Fila con encabezados reales
        if ("fecha" in row_text and 
            ("descripcion" in row_text or "descripción" in row_text) and 
            ("deposito" in row_text or "depositos" in row_text)):
            return i, False
    
    return None, False

# Función auxiliar para debug
def debug_mercado_pago_file(path):
    """
    Función solo para debug - muestra el contenido real del archivo
    """
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        print("=== DEBUG: CONTENIDO COMPLETO DEL ARCHIVO ===")
        print(content)
        print("="*50)
        
        # Guardar copia para análisis
        with open('debug_mercado_pago.txt', 'w', encoding='utf-8') as f:
            f.write(content)
        
        print("Archivo de debug guardado como: debug_mercado_pago.txt")
        
    except Exception as e:
        print(f"Error en debug: {e}")

def read_table_any(path, nrows=None):
    """
    Intenta leer Excel o CSV con varios separadores/encodings.
    Devuelve un DataFrame con columnas normalizadas en minúsculas.
    """
    df = None
    
    # 1) Excel primero
    try:
        df = pd.read_excel(path, nrows=nrows)
    except Exception:
        pass
    
    # 2) Si no, CSV con varios seps / encodings
    if df is None:
        for sep in [",", ";", "\t", "|"]:
            for enc in ["utf-8", "latin1", "cp1252"]:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc, nrows=nrows)
                    raise StopIteration
                except StopIteration:
                    break
                except Exception:
                    pass
            if df is not None:
                break
    
    if df is None:
        raise ValueError("No se pudo leer el archivo como Excel ni CSV.")
    
    _norm_cols(df)
    return df

def _norm_cols_no_parens(df):
    """
    Normaliza encabezados conservando el texto *antes* de paréntesis.
    'Depósitos (1)' -> 'depositos'
    """
    import re
    from unidecode import unidecode
    
    new_cols = []
    for c in df.columns:
        s = str(c)
        # quita todo lo que esté entre paréntesis
        s = re.sub(r"\(.*?\)", "", s)
        s = unidecode(s).strip().lower()
        s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")
        new_cols.append(s)
    
    df.columns = new_cols
    return df

def parse_banamex(path):
    """
    Lee TODAS las hojas y busca la tabla de movimientos.
    - Detecta encabezado embebido.
    - Toma solo depósitos > 0.
    - Devuelve filas listas para Transferencia.
    """
    try:
        # Intentar como Excel con todas las hojas
        dfs_dict = None
        try:
            dfs_dict = pd.read_excel(path, header=None, sheet_name=None)
        except Exception:
            dfs_dict = None
        
        # Si no es Excel, intentar como CSV (1 sola "hoja")
        if dfs_dict is None:
            try:
                df_csv = pd.read_csv(path, header=None, encoding="latin1")
                dfs_dict = {"__csv__": df_csv}
            except Exception:
                return []
        
        total_rows = []
        for sheet_name, df_raw in dfs_dict.items():
            if df_raw is None or df_raw.empty:
                continue
            
            hdr_idx, use_next = _find_header_row_banamex(df_raw)
            if hdr_idx is None:
                continue
            
            # Tomar fila de encabezados
            headers_row = hdr_idx + (1 if use_next else 0)
            if headers_row >= len(df_raw):
                continue
                
            headers = df_raw.iloc[headers_row].astype(str).tolist()
            df = df_raw.iloc[headers_row+1:].copy()
            df.columns = headers
            
            # Normaliza (quita paréntesis y acentos)
            _norm_cols_no_parens(df)
            
            # Elegir columnas claves con alias flexibles
            def pick_col(cands):
                for cand in cands:
                    cand = cand.strip().lower()
                    for col in df.columns:
                        if col == cand or cand in col:
                            return col
                return None
            
            col_fecha = pick_col(["fecha"])
            col_desc = pick_col(["descripcion","descripción"])
            col_dep = pick_col(["depositos","deposito"])
            col_fact = pick_col(["factura"])
            
            # Log útil para depurar
            app.logger.info(f"[Banamex] Hoja '{sheet_name}': columnas={list(df.columns)}")
            app.logger.info(f"[Banamex] Mapeo: fecha={col_fecha}, desc={col_desc}, dep={col_dep}, factura={col_fact}")
            
            if not (col_fecha and col_desc and col_dep):
                continue
                
            df = df[[c for c in [col_fecha, col_desc, col_dep, col_fact] if c in df.columns]].copy()
            df = df.dropna(how="all")
            
            # Construir filas
            for i, r in df.iterrows():
                monto = _to_float(r.get(col_dep))
                if monto <= 0:
                    continue
                    
                fecha = _to_date_any(r.get(col_fecha), dayfirst=True)
                if not fecha:
                    try:
                        fecha = pd.to_datetime(str(r.get(col_fecha)).strip(), dayfirst=True, errors='coerce')
                        fecha = fecha.date() if pd.notna(fecha) else None
                    except Exception:
                        fecha = None
                    if not fecha:
                        continue
                
                concepto = str(r.get(col_desc, "") or "").strip()
                factura = str(r.get(col_fact, "") or "").strip()
                
                ref = _mk_referencia_fallback("Banamex", fecha.isoformat(), monto, concepto, i)
                
                total_rows.append(dict(
                    fecha=fecha.isoformat(),
                    banco="Desconocido",
                    banco_receptor="Banamex",
                    monto=monto,
                    referencia=ref,
                    concepto=concepto,
                    pedido="",
                    factura=factura,
                ))
        
        return total_rows
        
    except Exception as e:
        app.logger.error(f"[Banamex] Error general: {e}")
        flash("No se pudo procesar el archivo de Banamex.", "error")
        return []

# --- Fallback: tu parseo "anterior" (reutiliza tu lógica de abonos/depósitos) ---
def _parse_banamex_fallback_old(path):
    try:
        try:
            df = pd.read_excel(path)
        except Exception:
            df = pd.read_csv(path, encoding='latin1', sep=',')
        
        _norm_cols(df)
        
        fecha_col = next((c for c in df.columns if 'fecha' in c), None)
        desc_col = next((c for c in df.columns if 'descripcion' in c or 'concepto' in c), None)
        deposito_col= next((c for c in df.columns if 'depositos' in c or 'abono' in c), None)
        factura_col = next((c for c in df.columns if 'factura' in c), None)
        
        if not fecha_col or not deposito_col:
            return []
        
        rows = []
        for i, r in df.iterrows():
            monto = _to_float(r.get(deposito_col))
            if monto <= 0:
                continue
                
            fecha = _to_date_any(r.get(fecha_col), dayfirst=True)
            if not fecha:
                continue
                
            concepto = str(r.get(desc_col, '')).strip() if desc_col else ''
            factura = str(r.get(factura_col, '')).strip() if factura_col else ''
            
            referencia = _mk_referencia_fallback("Banamex", fecha.isoformat(), monto, concepto, i)
            
            rows.append(dict(
                fecha=fecha.isoformat(),
                banco="Banamex",
                banco_receptor="Desconocido",
                monto=monto,
                referencia=referencia,
                concepto=concepto,
                pedido="",
                factura=factura,
            ))
        
        return rows
        
    except Exception:
        return []

def _find_header_row_bbva(df_raw, max_scan=80):
    """Devuelve el índice de la fila que parece contener los encabezados BBVA."""
    from unidecode import unidecode
    import re
    
    def norm_cell(v):
        s = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
        s = unidecode(s).lower().strip()
        s = re.sub(r"\s+", " ", s)
        return s
    
    nrows = min(len(df_raw), max_scan)
    for i in range(nrows):
        vals = [norm_cell(x) for x in df_raw.iloc[i].tolist()]
        row_text = " | ".join([v for v in vals if v])
        if not row_text:
            continue
            
        if (("dia" in row_text or "fecha" in row_text) and 
            ("concepto" in row_text or "referencia" in row_text) and 
            ("abono" in row_text or "deposito" in row_text or "depositos" in row_text or "importe" in row_text)):
            return i
    
    return None

def _infer_bbva_columns(df):
    """Inferencia por CONTENIDO para (fecha, concepto, abono, factura). No depende de los nombres de columna."""
    import numpy as np
    
    cols = list(df.columns)
    
    # 1) FECHA: mayor proporción de celdas que parsean como fecha
    best_fecha, best_fecha_score = None, -1.0
    for c in cols:
        s = df[c]
        total = min(len(s), 200)
        if total == 0:
            continue
        ok = 0
        for v in s.head(total):
            if _to_date_any(v, dayfirst=True):
                ok += 1
        score = ok / total
        if score > best_fecha_score:
            best_fecha, best_fecha_score = c, score

    # 2) ABONO: columna con más números POSITIVOS (tolera moneda/strings)
    best_abono, best_abono_score = None, -1.0
    for c in cols:
        s = df[c]
        total = min(len(s), 200)
        if total == 0:
            continue
        pos = 0
        for v in s.head(total):
            if _to_float(v) > 0:
                pos += 1
        score = pos / total
        if score > best_abono_score:
            best_abono, best_abono_score = c, score

    # 3) CONCEPTO: columna texto con mayor longitud media
    best_desc, best_len = None, -1.0
    for c in cols:
        s = df[c].astype(str).head(200).str.len()
        m = s.mean()
        if m > best_len:
            best_desc, best_len = c, m

    # 4) FACTURA (opcional): patrones tipo MA/LS 4xxxx
    best_fac, best_fac_hits = None, -1
    for c in cols:
        s = df[c].astype(str).head(200).str.upper()
        hits = s.str.contains(r"\b(MA|LS)\s*\d{3,6}\b", regex=True).sum()
        if hits > best_fac_hits:
            best_fac, best_fac_hits = c, int(hits)

    # Umbrales mínimos para considerar confiable
    if best_fecha_score < 0.25 or best_abono_score < 0.25:
        return None, None, None, None

    return best_fecha, best_desc, best_abono, (best_fac if best_fac_hits > 0 else None)


def parse_bbva(path):
    """Parser robusto para estados BBVA:
    - Solo ABONOS > 0
    - Normaliza encabezados (unidecode, lower, reemplaza '/')
    - Si la primera fila no son títulos, detecta fila de encabezados
    - Si no encuentra por nombre, infiere por contenido
    """
    try:
        import re
        
        def norm_hdr(s):
            s = unidecode(str(s)).strip().lower()
            s = re.sub(r"\s*/\s*", " ", s)  # "concepto / referencia" -> "concepto referencia"
            s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")
            return s

        def pick(df, *cands):
            for cand in cands:
                c = cand.strip().lower()
                for col in df.columns:
                    if col == c or c in col:
                        return col
            return None

        # ---- A) Leer como Excel; si no, intentar CSVs comunes
        df = None
        try:
            df = pd.read_excel(path)
        except Exception:
            pass

        if df is None or df.empty:
            for sep in [";", ",", "\t", "|"]:
                for enc in ["utf-8", "latin1", "cp1252"]:
                    try:
                        df = pd.read_csv(path, sep=sep, encoding=enc)
                        raise StopIteration
                    except StopIteration:
                        break
                    except Exception:
                        pass
                if df is not None and not df.empty:
                    break

        if df is None or df.empty:
            flash("No se pudo leer el archivo BBVA (Excel o CSV).", "error")
            return []

        df = df.dropna(how="all")
        df.columns = [norm_hdr(c) for c in df.columns]
        app.logger.info(f"[BBVA] cols(A/norm): {list(df.columns)}")

        col_fecha = pick(df, "dia", "fecha")
        col_desc = pick(df, "concepto_referencia", "concepto", "referencia", "descripcion")
        col_abono = pick(df, "abono", "abonos", "deposito", "depositos", "importe", "monto")
        col_factura = pick(df, "factura")

        # ---- B) Si no aparecen, buscar fila de encabezado en crudo
        if not (col_fecha and col_desc and col_abono):
            df_raw = None
            try:
                df_raw = pd.read_excel(path, header=None)
            except Exception:
                pass

            if df_raw is None or df_raw.empty:
                for sep in [";", ",", "\t", "|"]:
                    for enc in ["utf-8", "latin1", "cp1252"]:
                        try:
                            df_raw = pd.read_csv(path, header=None, sep=sep, encoding=enc)
                            raise StopIteration
                        except StopIteration:
                            break
                        except Exception:
                            pass
                    if df_raw is not None and not df_raw.empty:
                        break

            if df_raw is not None and not df_raw.empty:
                hdr = _find_header_row_bbva(df_raw)
                app.logger.info(f"[BBVA] header row: {hdr}")
                if hdr is not None:
                    headers = [norm_hdr(x) for x in df_raw.iloc[hdr].tolist()]
                    df = df_raw.iloc[hdr+1:].copy()
                    df.columns = headers
                    df = df.dropna(how="all")
                    app.logger.info(f"[BBVA] cols(B/norm): {list(df.columns)}")

                    col_fecha = pick(df, "dia", "fecha")
                    col_desc = pick(df, "concepto_referencia", "concepto", "referencia", "descripcion")
                    col_abono = pick(df, "abono", "abonos", "deposito", "depositos", "importe", "monto")
                    col_factura = pick(df, "factura")

        # ---- C) Heurística por contenido
        if not (col_fecha and col_desc and col_abono):
            f, d, a, fac = _infer_bbva_columns(df)
            col_fecha = col_fecha or f
            col_desc = col_desc or d
            col_abono = col_abono or a
            col_factura = col_factura or fac

        if not (col_fecha and col_desc and col_abono):
            app.logger.warning("[BBVA] No se hallaron columnas requeridas (dia/fecha, concepto, abono).")
            flash("El archivo no contiene columnas reconocibles de fecha/concepto/abono.", "error")
            return []

        # ---- D) Procesar SOLO abonos > 0
        rows, omitidas = [], 0
        for i, r in df.iterrows():
            monto = _to_float(r.get(col_abono))
            if monto <= 0:
                omitidas += 1
                continue

            fecha = _to_date_any(r.get(col_fecha), dayfirst=True)
            if not fecha:
                try:
                    dt = pd.to_datetime(str(r.get(col_fecha)).strip(), dayfirst=True, errors="coerce")
                    fecha = dt.date() if pd.notna(dt) else None
                except Exception:
                    fecha = None

            if not fecha:
                omitidas += 1
                continue

            concepto = str(r.get(col_desc, "") or "").strip()
            factura = str(r.get(col_factura, "") or "").strip() if col_factura else ""
            ref = _mk_referencia_fallback("BBVA", fecha.isoformat(), monto, concepto, i)

            rows.append(dict(
                fecha=fecha.isoformat(),
                banco="Desconocido",
                banco_receptor="BBVA",
                monto=monto,
                referencia=ref,
                concepto=concepto,
                pedido="",
                factura=factura,
            ))

        if not rows:
            app.logger.info(f"[BBVA] Sin abonos válidos. Omitidas: {omitidas}")
            flash("No se detectaron movimientos válidos en el archivo.", "warning")
            
        return rows

    except Exception as e:
        app.logger.error(f"[BBVA] Error general: {e}")
        flash("No se pudo procesar el archivo de BBVA.", "error")
        return []
    
@login_manager.user_loader
def load_user(user_id):
    try:
        return db.session.get(User, int(user_id))
    except Exception:
        # Si la DB no responde, evita que Flask-Login truene.
        return None

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('No tienes permisos para acceder a esta página', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_first_request
def initialize_database_if_needed():
    """Se ejecuta al primer request (por worker). Idempotente."""
    try:
        # Evita romper el primer request si la BD aún no responde
        if not db_ready():
            app.logger.warning("DB no lista aún en first_request; se intentará en el siguiente request.")
            return

        db.create_all()
        if User.query.count() == 0:
            admin = User(username='admin', is_admin=True)
            admin.password = os.getenv('ADMIN_PASSWORD', 'admin123')
            db.session.add(admin)
            db.session.commit()
            app.logger.info("Usuario admin creado automáticamente.")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error inicializando la BD: {e}")


# Rutas de autenticación mejoradas
@app.route('/', methods=['GET', 'POST'])
def login():
    # --- Healthcheck DB antes de tocar User.query ---
    if not db_ready():
        return "Base de datos no disponible, intenta en unos segundos.", 503

    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if User.query.count() == 0:
        return redirect(url_for('registro_usuario'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            flash('Por favor ingrese usuario y contraseña', 'error')
            return redirect(url_for('login'))
        
        user = User.query.filter_by(username=username).first()
        if user and user.verify_password(password):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        
        flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/registro-usuario', methods=['GET', 'POST'])
def registro_usuario():
    # Healthcheck antes de cualquier query
    if not db_ready():
        return "Base de datos no disponible, intenta en unos segundos.", 503

    # Si ya hay un usuario y no estás logeado, redirige a login
    if User.query.count() > 0 and not current_user.is_authenticated:
        flash("Ya existe un usuario registrado.")
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        confirm_password = request.form.get('confirm_password')

        if not username or not password:
            flash('Todos los campos son requeridos', 'error')
        elif password != confirm_password:
            flash('Las contraseñas no coinciden', 'error')
        elif User.query.filter_by(username=username).first():
            flash('El usuario ya existe', 'error')
        else:
            nuevo = User(username=username, is_admin=(User.query.count() == 0))
            nuevo.password = password
            db.session.add(nuevo)
            db.session.commit()
            flash('Usuario creado exitosamente. Ahora puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))

    return render_template('registro_usuario.html')

from flask import current_app, request

@app.before_request
def vv_fail_open_login():
    """
    Si la ruta es /vv (o subrutas) y la DB no está lista,
    desactiva temporalmente el requisito de login para esta petición.
    En el resto, se mantiene el login normal.
    """
    try:
        is_vv = request.path == "/vv" or request.path.startswith("/vv/")
        if is_vv and not db_ready():
            current_app.config["LOGIN_DISABLED"] = True
        else:
            current_app.config["LOGIN_DISABLED"] = False
    except Exception:
        # Si algo raro pasa aquí, mejor no desactivar login
        current_app.config["LOGIN_DISABLED"] = False


@app.route('/admin/usuarios')
@login_required
@admin_required
def admin_usuarios():
    usuarios = User.query.all()
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_usuario():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        confirm_password = request.form.get('confirm_password')
        is_admin = 'is_admin' in request.form
        
        if not username or not password:
            flash('Todos los campos son requeridos', 'error')
        elif password != confirm_password:
            flash('Las contraseñas no coinciden', 'error')
        elif User.query.filter_by(username=username).first():
            flash('El usuario ya existe', 'error')
        else:
            nuevo = User(username=username, is_admin=is_admin)
            nuevo.password = password
            db.session.add(nuevo)
            db.session.commit()
            flash('Usuario creado exitosamente.', 'success')
            return redirect(url_for('admin_usuarios'))
    
    return render_template('nuevo_usuario.html')

@app.route('/admin/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(id):
    usuario = User.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Debug: Mostrar datos recibidos
            app.logger.debug(f"Datos recibidos: {request.form}")
            
            # Validación de datos
            username = request.form.get('username', '').strip()
            if not username:
                flash('El nombre de usuario es requerido', 'error')
                return redirect(url_for('editar_usuario', id=id))
            
            # Verificar si el username ya existe
            if User.query.filter(User.username == username, User.id != id).first():
                flash('Este nombre de usuario ya está en uso', 'error')
                return redirect(url_for('editar_usuario', id=id))
            
            # Actualizar datos
            usuario.username = username
            usuario.is_admin = 'is_admin' in request.form
            
            # Manejar cambio de contraseña
            new_password = request.form.get('password', '').strip()
            if new_password:
                confirm_password = request.form.get('confirm_password', '').strip()
                if len(new_password) < 6:
                    flash('La contraseña debe tener al menos 6 caracteres', 'error')
                    return redirect(url_for('editar_usuario', id=id))
                if new_password != confirm_password:
                    flash('Las contraseñas no coinciden', 'error')
                    return redirect(url_for('editar_usuario', id=id))
                usuario.password = new_password
            
            db.session.commit()
            flash('Usuario actualizado correctamente', 'success')
            return redirect(url_for('admin_usuarios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el usuario: {str(e)}', 'error')
            app.logger.error(f"Error al editar usuario {id}: {str(e)}")
            return redirect(url_for('editar_usuario', id=id))
    
    return render_template('editar_usuario.html', usuario=usuario)

@app.route('/admin/usuarios/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_required
def eliminar_usuario(id):
    try:
        usuario = User.query.get_or_404(id)
        
        # Verificación de seguridad adicional
        if usuario.id == current_user.id:
            flash('No puedes eliminar tu propio usuario.', 'error')
            return redirect(url_for('admin_usuarios'))
        
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuario eliminado correctamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al eliminar usuario {id}: {str(e)}")
        flash('Ocurrió un error al eliminar el usuario.', 'error')
    
    return redirect(url_for('admin_usuarios'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    transferencia = Transferencia.query.get_or_404(id)
    
    if request.method == 'POST':
        transferencia.pedido = request.form['pedido']
        transferencia.factura = request.form['factura']
        transferencia.esta_registrado = 'esta_registrado' in request.form
        transferencia.registrado = request.form['registrado']
        db.session.commit()
        flash("Transferencia actualizada correctamente.")
        return redirect(url_for('dashboard'))
    
    return render_template('editar.html', t=transferencia)

@app.route('/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar(id):
    t = Transferencia.query.get_or_404(id)
    db.session.delete(t)
    db.session.commit()
    flash("Transferencia eliminada correctamente.")
    return redirect(url_for('dashboard'))

@app.route('/eliminar-todas', methods=['POST'])
@login_required
def eliminar_todas():
    Transferencia.query.delete()
    db.session.commit()
    flash('Se eliminaron todas las transferencias.')
    return redirect(url_for('dashboard'))

@app.route('/toggle-registrado/<int:id>', methods=['POST'])
@login_required
def toggle_registrado(id):
    t = Transferencia.query.get_or_404(id)
    t.esta_registrado = 'esta_registrado' in request.form
    db.session.commit()
    return redirect(url_for('dashboard'))

@app.route('/registro', methods=['GET', 'POST'])
@login_required
def registro():
    if request.method == 'POST':
        ref = request.form['referencia']
        existente = Transferencia.query.filter_by(referencia=ref).first()
        if existente:
            flash("Ya existe una transferencia con esa referencia.", 'error')
        else:
            t = Transferencia(
                fecha=request.form['fecha'],
                banco=request.form['banco'],
                monto=float(str(request.form['monto']).replace(',', '').strip() or 0),
                referencia=ref,
                pedido=request.form['pedido'],
                factura=request.form['factura'],
                registrado=request.form['registrado'],
                esta_registrado='esta_registrado' in request.form,
                concepto=request.form.get('concepto', '')
            )
            db.session.add(t)
            db.session.commit()
            flash("Transferencia registrada correctamente.", 'success')
            return redirect(url_for('dashboard'))
    
    # Pasa datetime al contexto de la plantilla
    return render_template('registro.html', datetime=datetime)

@app.route('/subir-archivo', methods=['GET', 'POST'])
@login_required
def subir_archivo():
    if request.method == 'POST':
        archivo = request.files['archivo']
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            archivo.save(filepath)
            
            try:
                if filename.endswith('.csv'):
                    df = pd.read_csv(filepath, dtype={'fecha': str})
                else:
                    df = pd.read_excel(filepath)
                
                # Normaliza encabezados
                df.columns = [col.strip().lower() for col in df.columns]
                
                # Renombrados útiles
                if 'banco participante' in df.columns:
                    df.rename(columns={'banco participante': 'banco_participante'}, inplace=True)
                
                # Fecha cruda
                if 'fecha' in df.columns:
                    df.rename(columns={'fecha': 'fecha_raw'}, inplace=True)
                elif 'fecha de movimiento' in df.columns:
                    df.rename(columns={'fecha de movimiento': 'fecha_raw'}, inplace=True)
                
                # Columna Cargo/Abono
                posibles_signos = [
                    'cargo/abono', 'cargo_abono', 'cargo-abono', 'cargo o abono',
                    'abono/cargo', 'signo', 'cargoabono'
                ]
                sign_col = next((c for c in posibles_signos if c in df.columns), None)
                if not sign_col:
                    flash("El archivo debe contener la columna 'Cargo/Abono' (por ejemplo: Cargo/Abono, cargo_abono, cargo-abono).", 'error')
                    return redirect(request.url)
                
                nuevas = 0
                duplicadas = 0
                saltadas_no_ingreso = 0
                omitidas_sin_referencia = 0
                
                for _, row in df.iterrows():
                    # 1) Solo ingresos
                    signo_raw = row.get(sign_col, '')
                    signo = str(signo_raw).strip().lower()
                    if signo not in ('+', 'abono'):
                        saltadas_no_ingreso += 1
                        continue
                    
                    # 2) Referencia
                    referencia = str(row.get('referencia', '')).strip()
                    if not referencia:
                        # Opción A: Omitir y contar
                        # omitidas_sin_referencia += 1
                        # continue
                        
                        # Opción B: Generar referencia estable para no perder la fila
                        fecha_tmp = str(row.get('fecha_raw', '')).strip()
                        concepto_tmp = str(row.get('concepto', '')).strip()
                        monto_tmp = _to_float(row.get('importe', 0) or row.get('monto', 0) or row.get('cantidad', 0))
                        referencia = _mk_referencia_fallback("EXTRACTO", fecha_tmp, monto_tmp, concepto_tmp, _)
                    
                    # 3) Deduplicar
                    if Transferencia.query.filter_by(referencia=referencia).first():
                        duplicadas += 1
                        continue
                    
                    # 4) Fecha robusta
                    raw_fecha = row.get('fecha_raw', '')
                    fecha = None
                    try:
                        if pd.isna(raw_fecha) or not str(raw_fecha).strip():
                            raise ValueError("Fecha vacía")
                        
                        if isinstance(raw_fecha, (int, float)):
                            raw_fecha = str(int(raw_fecha)).zfill(8)
                        else:
                            raw_fecha = str(raw_fecha).strip().replace("'", "").replace('"', '').replace('‘', '').replace('’', '')
                        
                        formatos = [
                            '%d%m%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d',
                            '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y', '%d-%b-%y', '%d-%B-%Y', '%m/%d/%y',
                        ]
                        
                        for fmt in formatos:
                            try:
                                fecha = pd.to_datetime(raw_fecha, format=fmt).date().isoformat()
                                break
                            except Exception:
                                continue
                        
                        if not fecha:
                            fecha = pd.to_datetime(raw_fecha).date().isoformat()
                    except Exception as e:
                        print(f"Error al convertir fecha: {raw_fecha} → {e}")
                        fecha = datetime.now().date().isoformat()
                    
                    # 5) Banco emisor
                    banco = str(row.get('banco_participante', '')).strip() or 'Desconocido'
                    
                    # 6) Banco receptor por cuenta/clabe
                    def _s(v):
                        s = '' if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
                        s = s.strip().replace("'", "")
                        return '' if s.lower() == 'nan' else s
                    
                    cuenta = _s(row.get('cuenta', ''))
                    clabe = _s(row.get('clabe', ''))
                    cuenta_completa = cuenta or clabe
                    banco_receptor = 'Desconocido'
                    
                    cuentas_destino = {
                        'BBVA': ['0185077915', '012420001850779158'],
                        'BANAMEX': ['53700061612', '002420053700616125'],
                        'SANTANDER': ['5150036891', '014420515003689123', '51500368912']
                    }
                    
                    for banco_nombre, cuentas in cuentas_destino.items():
                        for c in cuentas:
                            if cuenta_completa.endswith(c) or cuenta_completa == c:
                                banco_receptor = banco_nombre
                                break
                        if banco_receptor != 'Desconocido':
                            break
                    
                    # 7) Monto y campos
                    monto = _to_float(row.get('importe', 0) or row.get('monto', 0) or row.get('cantidad', 0))
                    pedido = str(row.get('pedido', '')).strip()
                    factura = str(row.get('factura', '')).strip()
                    concepto = str(row.get('concepto', '')).strip()
                    
                    # 8) Guardar
                    t = Transferencia(
                        fecha=fecha,
                        banco=banco,
                        banco_receptor=banco_receptor,
                        monto=monto,
                        referencia=referencia,
                        pedido=pedido,
                        factura=factura,
                        registrado=current_user.username,
                        esta_registrado=False,
                        concepto=concepto
                    )
                    db.session.add(t)
                    nuevas += 1
                
                db.session.commit()
                flash(
                    f"Se procesó el archivo. Nuevas: {nuevas}, "
                    f"Duplicadas (misma referencia): {duplicadas}, "
                    f"Omitidas por no ser ingreso (+/ABONO): {saltadas_no_ingreso}, "
                    f"Omitidas/generadas por falta de referencia: {omitidas_sin_referencia}", 'success'
                )
                return redirect(url_for('dashboard'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Ocurrió un error al procesar el archivo: {e}', 'error')
                return redirect(request.url)
        else:
            flash('Archivo no válido. Solo se permiten .csv, .xlsx o .xls', 'error')
    
    return render_template('subir_archivo.html')

@app.route('/facturas/aplicar', methods=['POST'], endpoint='facturas_aplicar_a_transferencias')
@login_required
def facturas_aplicar_a_transferencias():
    """
    Sube el Excel de 'Facturas Emitidas' y, por cada fila:
    - Construye la clave de factura 'Serie+Folio'
    - Intenta encontrar una transferencia que coincida por (fecha + monto) o por texto
    - Si hay match único -> set factura=<Serie+Folio>, esta_registrado=True, registrado=<usuario>
    """
    import pandas as pd
    
    archivo = request.files.get('archivo_emitidas')
    if not archivo:
        flash("No se subió el archivo de Facturas Emitidas.", "error")
        return redirect(url_for('dashboard') + '#tab-transferencias')
    
    # ---------- Helpers locales ----------
    def _norm_cols_emitidas(df):
        import re
        from unidecode import unidecode
        
        new_cols = []
        for c in df.columns:
            s = str(c)
            s = re.sub(r"\(.*?\)", "", s)  # quita paréntesis
            s = unidecode(s).strip().lower()
            s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")
            new_cols.append(s)
        df.columns = new_cols
        return df
    
    def pick(df, *names):
        cols = list(df.columns)
        for raw in names:
            key = raw.strip().lower()
            # match exacto
            for c in cols:
                if c == key:
                    return c
            # match por contiene
            for c in cols:
                if key in c:
                    return c
        return None
    
    def s_(x):
        s = '' if x is None else str(x).strip()
        return '' if s.lower() in ('nan', 'none') else s
    
    def f_(x):
        try:
            return _to_float(x)
        except Exception:
            return 0.0
    
    def d_(x):
        try:
            return _to_date_any(x, dayfirst=True)
        except Exception:
            return None
    
    # ---------- Leer Excel ----------
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        flash(f"No se pudo leer el Excel de Emitidas: {e}", "error")
        return redirect(url_for('dashboard') + '#tab-transferencias')
    
    if df.empty:
        flash("El archivo no contiene filas.", "warning")
        return redirect(url_for('dashboard') + '#tab-transferencias')
    
    _norm_cols_emitidas(df)
    
    # Campos típicos a detectar
    col_serie = pick(df, "serie")
    col_folio = pick(df, "folio")
    col_tipo = pick(df, "tipo", "tipo_de_comprobante", "tipodocumento")
    col_total = pick(df, "total", "importe_total", "total_factura", "monto_total", "importe")
    col_fecha = pick(df, "fecha", "fecha_emision", "fecha_de_emision", "fecha_factura")
    
    if not (col_serie and col_folio):
        flash("No se localizaron columnas 'Serie' y 'Folio' en el archivo de Emitidas.", "error")
        return redirect(url_for('dashboard') + '#tab-transferencias')
    
    # Stats
    procesadas = 0
    actualizadas = 0
    sin_match = 0
    ambiguas = 0
    ignoradas_sin_datos = 0
    
    # ---------- Procesar filas ----------
    for _, r in df.iterrows():
        serie = s_(r.get(col_serie))
        folio = s_(r.get(col_folio))
        if not serie and not folio:
            ignoradas_sin_datos += 1
            continue
        
        clave = f"{serie}{folio}".strip()
        # (Opcional) normalizar la clave por si trae espacios o caracteres raros
        # clave = _norm_factura(clave)
        if not clave:
            ignoradas_sin_datos += 1
            continue
        
        tipo = (s_(r.get(col_tipo)) or "").upper()
        total = f_(r.get(col_total, 0))
        fecha = d_(r.get(col_fecha))
        
        # Si no hay total o fecha, aún así intentamos por texto, pero lo contamos aparte
        match_hecho = False
        
        # 1) Intento por fecha + monto (si tenemos ambos)
        if fecha and (total != 0):
            # En Transferencia.fecha guardas string YYYY-MM-DD
            fecha_iso = fecha.isoformat()
            # Monto con tolerancia ±1.00
            cand1 = (Transferencia.monto >= (total - 1.0))
            cand2 = (Transferencia.monto <= (total + 1.0))
            candidatos = (Transferencia.query
                .filter(Transferencia.fecha == fecha_iso)
                .filter(cand1, cand2)
                .all())
            
            if len(candidatos) == 1:
                t = candidatos[0]
                t.factura = clave
                t.esta_registrado = True
                t.registrado = current_user.username
                actualizadas += 1
                match_hecho = True
            elif len(candidatos) > 1:
                # Si varios con mismo monto/fecha, probamos afinar por texto (case-insensitive)
                sub = [
                    tt for tt in candidatos
                    if (clave.lower() in (tt.referencia or "").lower()) or
                    (clave.lower() in (tt.pedido or "").lower()) or
                    (clave.lower() in (tt.concepto or "").lower()) or
                    (clave.lower() in (tt.factura or "").lower())
                ]
                if len(sub) == 1:
                    t = sub[0]
                    t.factura = clave
                    t.esta_registrado = True
                    t.registrado = current_user.username
                    actualizadas += 1
                    match_hecho = True
                elif len(sub) > 1:
                    ambiguas += 1
                    match_hecho = True  # considerado analizado aunque ambigua
        
        # 2) Si no hubo match por monto/fecha, intentamos por texto directamente (case-insensitive)
        if not match_hecho:
            q = (Transferencia.query.filter(
                or_(
                    Transferencia.referencia.ilike(f"%{clave}%"),
                    Transferencia.pedido.ilike(f"%{clave}%"),
                    Transferencia.concepto.ilike(f"%{clave}%"),
                    Transferencia.factura.ilike(f"%{clave}%")
                )
            ).all())
            
            if len(q) == 1:
                t = q[0]
                t.factura = clave
                t.esta_registrado = True
                t.registrado = current_user.username
                actualizadas += 1
            elif len(q) > 1:
                ambiguas += 1
            else:
                sin_match += 1
        
        procesadas += 1
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error guardando cambios: {e}", "error")
        return redirect(url_for('dashboard') + '#tab-transferencias')
    
    flash(
        f"Emitidas procesadas: {procesadas}. "
        f"Actualizadas: {actualizadas}. "
        f"Sin match: {sin_match}. "
        f"Ambiguas: {ambiguas}. "
        f"Ignoradas sin datos: {ignoradas_sin_datos}.", "success"
    )
    return redirect(url_for('dashboard') + '#tab-transferencias')

@app.route('/subir_bancos', methods=['POST'])
@login_required
def subir_bancos():
    """
    Sube extractos de Mercado Pago / Banamex / BBVA y:
    - crea nuevas transferencias
    - si la referencia ya existe => NO ACTUALIZA y la cuenta como 'duplicada sin cambios'
    - ignora movimientos rechazados ([RECHAZADO]...) y montos <= 0
    
    Además:
    - Hace un 'peek' de diagnóstico (columnas y primeras filas) para ver qué se leyó
    - Detección más robusta por columnas cuando el nombre del archivo no ayuda
    """
    archivo = request.files.get('archivo_banco')
    if not archivo or not allowed_file(archivo.filename):
        flash("Sube un archivo válido (.xls, .xlsx, .csv).", "error")
        return redirect(url_for('dashboard'))
    
    # Guarda temporal
    filename = secure_filename(archivo.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    archivo.save(path)
    fname = filename.lower()
    filas = []
    
    # --- DIAGNÓSTICO: intenta ver columnas y primeras filas ---
    try:
        dfpeek = None
        origen = "desconocido"
        try:
            dfpeek = pd.read_excel(path, nrows=8)
            origen = "excel"
        except Exception:
            # intenta CSV con varios separadores y encodings
            for sep in [",", ";", "\t", "|"]:
                for enc in ["utf-8", "latin1", "cp1252"]:
                    try:
                        dfpeek = pd.read_csv(path, nrows=8, sep=sep, encoding=enc)
                        origen = f"csv sep={sep} enc={enc}"
                        break  # Cambiado de raise StopIteration a break
                    except Exception:
                        pass
                if dfpeek is not None:
                    break
        
        if dfpeek is not None:
            # Normaliza nombres para imprimirlos coherentes
            cols_norm = [str(c).strip().lower() for c in dfpeek.columns]
            app.logger.info(f"[subir_bancos] Archivo: {filename}")
            app.logger.info(f"[subir_bancos] Origen detectado: {origen}")
            app.logger.info(f"[subir_bancos] Columnas detectadas: {cols_norm}")
            try:
                app.logger.info(f"[subir_bancos] Head:\n{dfpeek.head().to_string(index=False)}")
            except Exception:
                app.logger.info("[subir_bancos] No se pudo imprimir head()")
    except Exception as _e:
        app.logger.warning(f"[subir_bancos] Peek falló: {_e}")
    
    try:
        # 1) Detección por nombre de archivo
        if 'mercado' in fname or 'mp' in fname:
            filas = parse_mercado_pago(path)
        elif 'banamex' in fname or 'citibanamex' in fname or 'bmx' in fname:
            filas = parse_banamex(path)
        elif 'bbva' in fname:
            filas = parse_bbva(path)
        else:
            # 2) Fallback por columnas/contenido (sin asumir formato)
            # Relee de forma robusta SOLO para detectar tipo
            df = None
            try:
                df = pd.read_excel(path, nrows=3)
            except Exception:
                for sep in [",", ";", "\t", "|"]:
                    for enc in ["utf-8", "latin1", "cp1252"]:
                        try:
                            df = pd.read_csv(path, nrows=3, sep=sep, encoding=enc)
                            break  # Cambiado de raise StopIteration a break
                        except Exception:
                            pass
                    if df is not None:
                        break
            
            if df is None or df.empty:
                filas = []
            else:
                cs = [str(c).strip().lower() for c in df.columns]
                texto_cols = " ".join(cs)
                # Pistas típicas por contenido de columnas
                if any(k in texto_cols for k in ['mercado', 'mercado pago', 'mp']):
                    filas = parse_mercado_pago(path)
                elif any(k in texto_cols for k in ['banamex', 'citi', 'citibanamex']):
                    filas = parse_banamex(path)
                elif any(k in texto_cols for k in ['bbva', 'retiro', 'abono', 'cargo']):
                    filas = parse_bbva(path)
                else:
                    # Último intento: Mercado Pago (suele traer fecha/monto estándar)
                    filas = parse_mercado_pago(path)
        
        if not filas:
            flash("No se detectaron movimientos válidos en el archivo.", "warning")
            return redirect(url_for('dashboard'))
        
        nuevas, duplicadas = 0, 0
        ignoradas_rechazado, ignoradas_monto = 0, 0
        
        def es_rechazado(texto):
            s = (str(texto or '')).strip().lower()
            return s.startswith('[rechazado]') or 'rechazado' in s
        
        for i, it in enumerate(filas):
            concepto_txt = it.get('concepto', '')
            if es_rechazado(concepto_txt):
                ignoradas_rechazado += 1
                continue
            
            monto = _to_float(it.get('monto', 0))
            if monto <= 0:
                ignoradas_monto += 1
                continue
            
            ref = (it.get('referencia', '') or '').strip()
            if not ref:
                # seguridad extra si el parser no generó referencia
                ref = _mk_referencia_fallback(
                    it.get('banco', 'DESCONOCIDO'),
                    it.get('fecha', ''),
                    monto,
                    concepto_txt,
                    i
                )
            
            existe = Transferencia.query.filter_by(referencia=ref).first()
            if existe:
                duplicadas += 1
                continue
            
            t = Transferencia(
                fecha=it.get('fecha', ''),
                banco=it.get('banco', 'Desconocido'),
                banco_receptor=(it.get('banco_receptor') or "Desconocido"),
                monto=monto,
                referencia=ref,
                pedido=it.get('pedido', ''),
                factura=it.get('factura', ''),
                registrado=current_user.username,
                esta_registrado=False,
                concepto=concepto_txt
            )
            db.session.add(t)
            nuevas += 1
        
        db.session.commit()
        flash(
            f"Extracto procesado. Nuevas: {nuevas}. "
            f"Duplicadas sin cambios: {duplicadas}. "
            f"Ignoradas rechazadas: {ignoradas_rechazado}. "
            f"Ignoradas monto<=0: {ignoradas_monto}.", "success"
        )
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error al procesar el extracto: {e}", "error")
    
    return redirect(url_for('dashboard'))

@app.route('/transformar_excel', methods=['POST'])
@login_required
def transformar_excel():
    archivo = request.files.get('archivo_excel')
    if not archivo:
        flash("No se subió ningún archivo", "error")
        return redirect(url_for('dashboard'))
    
    df = pd.read_excel(archivo)
    df.columns = [col.strip() for col in df.columns]
    
    columnas_requeridas = [
        "Cve_factu", "No_fac", "Num", "Fecha", "Cantidad", "Tipo",
        "No_nota", "Subtipo", "Cant", "Cve_age", "Nom_cte", "Rfc_cte", "Des_mon"
    ]
    
    if not all(col in df.columns for col in columnas_requeridas):
        flash("El archivo no contiene todas las columnas requeridas", "error")
        return redirect(url_for('dashboard'))
    
    # Convertir fecha a solo la parte de fecha (sin hora)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors='coerce').dt.date
    
    # Crear columna combinada sin espacios ni encabezado
    df.insert(0, '', df['Cve_factu'].astype(str).str.strip() + df['No_fac'].astype(str).str.strip())
    
    # Limpiar campos para evitar errores por espacios
    columnas_a_limpieza = ['Nom_cte', 'Rfc_cte', 'Des_mon']
    for col in columnas_a_limpieza:
        df[col] = df[col].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
    
    # Reordenar columnas
    columnas_orden = [
        '', 'Num', 'No_fac', 'Fecha', 'Cantidad', 'Tipo', 'No_nota',
        'Subtipo', 'Cant', 'Cve_age', 'Nom_cte', 'Rfc_cte', 'Des_mon'
    ]
    df_final = df[columnas_orden].copy()
    
    # Filtrar registros con "DEPOSITO CLI" en Subtipo
    df_final = df_final[~df_final["Subtipo"].astype(str).str.contains("DEPOSITO CLI", na=False)]
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    
    # Encabezados (con columna inicial sin título)
    encabezados = [
        "", "Num", "No_fac", "Fecha", "Cantidad", "Tipo", "No_nota",
        "Subtipo", "Cant", "Cve_age", "Nom_cte", "Rfc_cte", "Des_mon"
    ]
    ws.append(encabezados)
    
    # Agregar los datos sin encabezado automático
    for row in dataframe_to_rows(df_final, index=False, header=False):
        ws.append(row)
    
    # Estilo de encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Forzar primera columna a formato texto
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = '@'
    
    # Guardar archivo con timestamp (en UPLOAD_FOLDER)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"excel_transformado_{timestamp}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(output_path)
    
    return send_file(output_path, as_attachment=True)


@app.route('/transformar_excel_clasificado', methods=['POST'])
@login_required
def transformar_excel_clasificado():
    # 1) Recibir archivo con nombres alternativos
    archivo = (request.files.get('archivo_excel')
               or request.files.get('archivo')
               or request.files.get('file'))
    if not archivo or archivo.filename.strip() == "":
        flash("No se subió ningún archivo. Revisa que el <form> tenga enctype='multipart/form-data' y el input se llame 'archivo_excel'.", "error")
        app.logger.warning(f"[clasificado] request.files keys: {list(request.files.keys())}")
        return redirect(url_for('dashboard'))

    # 2) Leer Excel
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        flash(f"No pude leer el Excel: {e}", "error")
        return redirect(url_for('dashboard'))

    # 3) Normalizar encabezados (sin acentos, sin espacios raros)
    from unidecode import unidecode
    def norm_col(c):
        s = unidecode(str(c)).replace("\xa0", " ").strip()
        return s
    df.columns = [norm_col(c) for c in df.columns]

    # 4) Verificar / mapear columnas requeridas de forma tolerante
    requeridas = ["Cve_factu", "No_fac", "Num", "Fecha", "Cantidad", "Tipo",
                  "No_nota", "Subtipo", "Cant", "Cve_age", "Nom_cte", "Rfc_cte", "Des_mon"]

    # índice auxiliar para ignorar espacios y mayúsculas
    idx = {unidecode(str(c)).lower().replace(" ", ""): c for c in df.columns}
    mapeo = {}
    faltan = []
    for r in requeridas:
        key = unidecode(r).lower().replace(" ", "")
        col_real = idx.get(key)
        if col_real is None:
            faltan.append(r)
        else:
            mapeo[r] = col_real

    if faltan:
        flash("El archivo no contiene todas las columnas requeridas: " + ", ".join(faltan), "error")
        app.logger.warning(f"[clasificado] columnas detectadas: {list(df.columns)} | faltan: {faltan}")
        return redirect(url_for('dashboard'))

    # Renombrar a nombres “canónicos”
    df = df.rename(columns=mapeo)

    # ---------- Transformación (idéntica a la tuya, con pequeños fixes) ----------
    df.insert(0, '', df['Cve_factu'].astype(str).str.strip() + df['No_fac'].astype(str).str.strip())

    for col in ['Nom_cte', 'Rfc_cte', 'Des_mon']:
        df[col] = df[col].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()

    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce').dt.strftime('%d/%m/%Y')

    def clasificar(row):
        subtipo = str(row.get('Subtipo', '')).upper()
        if "CAJA 5MAYO" in subtipo:
            return "CAJA 5 MAYO"
        elif "SLORENZ" in subtipo:
            return "SAN LORENZO"
        elif "TEC" in subtipo:
            return "TECNOLÓGICO"
        elif "PENIN" in subtipo:
            return "PENINSULA"
        else:
            return "OTROS MOVIMIENTOS"

    df['__clasificacion_temp'] = df.apply(clasificar, axis=1)

    columnas_orden = [
        '', 'Num', 'No_fac', 'Fecha', 'Cantidad', 'Tipo', 'No_nota',
        'Subtipo', 'Cant', 'Cve_age', 'Nom_cte', 'Rfc_cte', 'Des_mon'
    ]
    df_final = df[columnas_orden + ['__clasificacion_temp']].copy()
    df_final = df_final[~df_final["Subtipo"].astype(str).str.contains("DEPOSITO CLI", na=False)]

    df_final["Cant"] = pd.to_numeric(df_final["Cant"], errors='coerce').fillna(0)
    df_final["Cantidad"] = pd.to_numeric(df_final["Cantidad"], errors='coerce').fillna(0)
    df_final["Tipo"] = df_final["Tipo"].astype(str).str.strip().str.upper()

    total_efectivo = df_final[df_final["Tipo"] == "EFECTIVO"]["Cantidad"].sum()

    subtotales = {}
    for s in ["CAJA 5 MAYO", "SAN LORENZO", "TECNOLÓGICO", "PENINSULA"]:
        subtotales[s] = df_final[df_final["__clasificacion_temp"] == s]["Cant"].sum()

    # ---------- Generar Excel en memoria (sin escribir a disco) ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificado"

    # Encabezado
    ws.append(columnas_orden)

    # Filas
    for row in dataframe_to_rows(df_final[columnas_orden], index=False, header=False):
        ws.append(row)

    # Estilo encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar anchos
    for col in ws.columns:
        max_length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

    # Fila resumen con longitud igual al encabezado
    fila_resumen = [""] * len(columnas_orden)
    fila_resumen[0] = f"TOTAL EFECTIVO: ${total_efectivo:,.2f}"
    # (Opcional) podrías agregar más totales en otras celdas si lo necesitas
    ws.append(fila_resumen)

    from openpyxl.styles import PatternFill
    azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for cell in ws[ws.max_row]:
        cell.fill = azul
        cell.font = Font(bold=True)

    # Descargar
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"excel_clasificado_{ts}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/dashboard')
@login_required
def dashboard():
    # ---------- Transferencias ----------
    t_query = Transferencia.query

    referencia = request.args.get('referencia', type=str)
    if referencia:
        t_query = t_query.filter(Transferencia.referencia.contains(referencia))

    # filtro de fecha para TRANSFERENCIAS (string yyyy-mm-dd)
    fecha = request.args.get('fecha', type=str)
    if fecha:
        t_query = t_query.filter(Transferencia.fecha == fecha)

    # banco receptor activo (para filtrar tabla y resaltar botón)
    banco_receptor_activo = request.args.get('banco_receptor', type=str)
    if banco_receptor_activo:
        t_query = t_query.filter(Transferencia.banco_receptor == banco_receptor_activo)

    transferencias = t_query.order_by(Transferencia.fecha.desc()).all()

    # ✅ Conteos por BANCO RECEPTOR para los botones
    rows_cont = db.session.query(
        Transferencia.banco_receptor,
        func.count(Transferencia.id)
    ).group_by(Transferencia.banco_receptor).all()

    bancos_contadores = {}
    for b, c in rows_cont:
        etiqueta = b if (b and str(b).strip()) else "Desconocido"
        bancos_contadores[etiqueta] = bancos_contadores.get(etiqueta, 0) + c

    total_transferencias = sum(bancos_contadores.values())

    # Orden: mayor conteo primero, luego alfabético
    bancos_ordenados = sorted(bancos_contadores.items(), key=lambda x: (-x[1], x[0]))

    # ---------- Ventas ----------
    ventas_query = Venta.query
    fecha_ventas = request.args.get('fecha_ventas', type=str)  # viene del form de la pestaña Ventas
    if fecha_ventas:
        try:
            f = datetime.strptime(fecha_ventas, "%Y-%m-%d").date()
            ventas_query = ventas_query.filter(Venta.fecha == f)
        except ValueError:
            flash("Fecha de ventas inválida. Usa el formato YYYY-MM-DD.", "error")

    ventas = ventas_query.order_by(Venta.fecha.desc()).all()

    # Clave combinada que usas en la tabla (por si la necesitas)
    for v in ventas:
        v.c = f"{(v.tipo or '')[:2].upper()}{(v.no_fac or '')}"

    return render_template(
        'dashboard.html',
        name=current_user.username,
        transferencias=transferencias,
        ventas=ventas,
        banco_receptor_activo=banco_receptor_activo,
        bancos_ordenados=bancos_ordenados,
        total_transferencias=total_transferencias
    )


@app.route('/subir_ventas', methods=['POST'])
@login_required
def subir_ventas():
    from sqlalchemy import func
    import math
    import numpy as np
    from datetime import date, datetime
    
    def clean_float(x, default=0.0):
        """Convierte a float seguro (sin NaN/inf ni tipos numpy)."""
        try:
            # Primero saca valores 'nan' en string
            if isinstance(x, str) and x.strip().lower() in ('nan', 'none', ''):
                return float(default)
            
            # Convierte numpy -> python
            if isinstance(x, (np.floating,)):
                x = float(x)
            
            v = float(x)
            if math.isnan(v) or math.isinf(v):
                return float(default)
            return v
        except Exception:
            return float(default)
    
    def clean_str(x):
        """Convierte a string seguro, evitando 'nan' literales."""
        s = '' if x is None else str(x).strip()
        return '' if s.lower() in ('nan', 'none') else s
    
    def clean_date(x):
        """Devuelve un datetime.date or None."""
        if x is None:
            return None
        if isinstance(x, date) and not isinstance(x, datetime):
            return x
        if isinstance(x, datetime):
            return x.date()
        
        # pandas / numpy / string
        try:
            # intenta varios formatos comunes
            return pd.to_datetime(x, dayfirst=True, errors='coerce').date()
        except Exception:
            return None
    
    archivo = request.files.get('archivo_excel')
    if not archivo:
        flash("No se subió ningún archivo", "error")
        return redirect(url_for('dashboard') + '#tab-ventas')
    
    filename = (archivo.filename or "").lower()
    
    # --- Caso especial: Mercado Pago ---
    if 'mercado' in filename or 'mp' in filename:
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(archivo.filename))
        archivo.save(temp_path)
        ventas_data = parse_mercado_pago_para_ventas(temp_path)
        
        if not ventas_data:
            flash("No se pudieron extraer ventas del archivo de Mercado Pago.", "error")
            return redirect(url_for('dashboard') + '#tab-ventas')
        
        nuevos = 0
        for row in ventas_data:
            try:
                row_cantidad = clean_float(row.get('cantidad', 0))
                row_cant = clean_float(row.get('cant', row_cantidad))
                fecha_row = clean_date(row.get('fecha'))
                if not fecha_row:
                    continue
                
                codigo_row = clean_str(row.get('codigo', ''))
                
                # Chequeo duplicado
                try:
                    dup = (db.session.query(func.count(Venta.id))
                        .filter(Venta.fecha == fecha_row, Venta.codigo == codigo_row, Venta.cantidad == row_cantidad)
                        .scalar() or 0)
                except Exception as e:
                    app.logger.warning(f"[VENTAS MP] fallo COUNT dup: {e}")
                    dup = 0
                
                if dup > 0:
                    continue
                
                row['fecha'] = fecha_row
                row['cantidad'] = row_cantidad
                row['cant'] = row_cant
                row['usuario'] = current_user.username
                
                campos_validos = {k: v for k, v in row.items() if hasattr(Venta, k)}
                db.session.add(Venta(**campos_validos))
                nuevos += 1
            except Exception as e:
                app.logger.error(f"[MP->Venta] Error guardando fila: {e}")
                continue
        
        db.session.commit()
        flash(f"Se procesaron {nuevos} ventas desde Mercado Pago.", "success")
        return redirect(url_for('dashboard') + '#tab-ventas')
    
    # --- Procesamiento normal (Excel sin encabezados) ---
    try:
        df = pd.read_excel(archivo, header=None)
    except Exception as e:
        flash(f"No se pudo leer el Excel de ventas: {e}", "error")
        return redirect(url_for('dashboard') + '#tab-ventas')
    
    columnas = [
        'codigo', 'Num', 'No_fac', 'Fecha', 'Cantidad', 'Tipo', 'No_nota', 'Subtipo', 'Cant', 'Cve_age', 'Nom_cte', 'Rfc_cte', 'Des_mon'
    ]
    
    if df.shape[1] < len(columnas):
        flash("El archivo de ventas no tiene el número de columnas esperado.", "error")
        return redirect(url_for('dashboard') + '#tab-ventas')
    
    df = df.iloc[:, :len(columnas)].copy()
    df.columns = columnas
    
    nuevos = 0
    for _, fila in df.iterrows():
        # Fecha
        fecha = clean_date(fila.get('Fecha'))
        if not fecha:
            continue
        
        # Montos (ambos campos)
        cantidad_val = clean_float(fila.get('Cantidad', 0), 0.0)
        cant_val = clean_float(fila.get('Cant', cantidad_val), cantidad_val)
        if cantidad_val == 0 and cant_val == 0:
            continue
        
        # Strings
        no_fac_norm = clean_str(fila.get('No_fac', ''))
        codigo_norm = clean_str(fila.get('codigo', ''))
        num_norm = clean_str(fila.get('Num', ''))
        no_nota_norm = clean_str(fila.get('No_nota', ''))
        tipo_norm = clean_str(fila.get('Tipo', ''))
        subtipo_norm = clean_str(fila.get('Subtipo', ''))
        cve_age_norm = clean_str(fila.get('Cve_age', ''))
        nom_cte_norm = clean_str(fila.get('Nom_cte', ''))
        rfc_cte_norm = clean_str(fila.get('Rfc_cte', ''))
        des_mon_norm = clean_str(fila.get('Des_mon', ''))
        
        # Chequeo duplicado seguro (sin NaN en binds)
        try:
            dup = (db.session.query(func.count(Venta.id))
                .filter(Venta.fecha == fecha, Venta.no_fac == no_fac_norm, Venta.cantidad == cantidad_val)
                .scalar() or 0)
        except Exception as e:
            app.logger.warning(f"[VENTAS XLS] fallo COUNT dup: {e} | fecha={fecha} no_fac='{no_fac_norm}' cantidad={cantidad_val}")
            dup = 0
        
        if dup > 0:
            continue
        
        venta = Venta(
            fecha=fecha,
            concepto=codigo_norm,
            tipo=tipo_norm,
            subtipo=subtipo_norm,
            cantidad=cantidad_val,
            usuario=current_user.username,
            codigo=codigo_norm,
            num=num_norm,
            no_fac=no_fac_norm,
            no_nota=no_nota_norm,
            cant=cant_val,
            cve_age=cve_age_norm,
            nom_cte=nom_cte_norm,
            rfc_cte=rfc_cte_norm,
            des_mon=des_mon_norm
        )
        db.session.add(venta)
        nuevos += 1
    
    db.session.commit()
    flash(f"Archivo de ventas cargado. Nuevos: {nuevos}", "success")
    return redirect(url_for('dashboard') + '#tab-ventas')

@app.route("/ventas/upload_clasificado", methods=["POST"])
@login_required
def ventas_upload_clasificado():
    archivo = request.files.get("archivo_excel")
    if not archivo:
        flash("No se subió ningún archivo", "error")
        return redirect(url_for("dashboard") + "#tab-ventas")
    
    try:
        # El clasificado que generas trae encabezados
        df = pd.read_excel(archivo)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Columnas que esperamos del clasificado (las que generas en transformar_excel_clasificado)
        esperadas = [
            "", "Num", "No_fac", "Fecha", "Cantidad", "Tipo", "No_nota", "Subtipo", "Cant", "Cve_age", "Nom_cte", "Rfc_cte", "Des_mon"
        ]
        
        faltantes = [c for c in esperadas if c not in df.columns]
        if faltantes:
            flash(f"Faltan columnas en el archivo clasificado: {', '.join(faltantes)}", "error")
            return redirect(url_for("dashboard") + "#tab-ventas")
        
        # Normaliza fecha (tu clasificado sale dd/mm/YYYY)
        def parse_fecha(v):
            if pd.isna(v):
                return None
            if isinstance(v, datetime):
                return v.date()
            s = str(v).strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%Y%m%d", "%d%m%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    continue
            try:
                return pd.to_datetime(s).date()
            except Exception:
                return None
        
        df["__fecha"] = df["Fecha"].apply(parse_fecha)
        
        nuevos = 0
        objetos = []
        for _, r in df.iterrows():
            fecha = r["__fecha"]
            if not fecha:
                continue
            
            # Montos
            cantidad_val = _to_float(r.get("Cantidad", 0) or 0)
            cant_val = _to_float(r.get("Cant", cantidad_val))
            
            # Si ambos son 0, no registrar
            if cantidad_val == 0 and cant_val == 0:
                continue
            
            # Primera columna combinada (sin nombre) la usamos como código y concepto
            codigo_val = str(r.get("", "")).strip()
            tipo_val = str(r.get("Tipo", "")).strip()
            subtipo_val = str(r.get("Subtipo", "")).strip()
            no_fac_val = str(r.get("No_fac", "")).strip()
            no_nota_val = str(r.get("No_nota", "")).strip()
            cve_age_val = str(r.get("Cve_age", "")).strip()
            nom_cte_val = str(r.get("Nom_cte", "")).strip()
            rfc_cte_val = str(r.get("Rfc_cte", "")).strip()
            des_mon_val = str(r.get("Des_mon", "")).strip()
            
            venta = Venta(
                fecha=fecha,
                concepto=codigo_val,
                tipo=tipo_val,
                subtipo=subtipo_val,
                cantidad=cantidad_val,  # ← desde "Cantidad"
                usuario=current_user.username,
                codigo=codigo_val,
                num=str(r.get("Num", "")).strip(),
                no_fac=no_fac_val,
                no_nota=no_nota_val,
                cant=cant_val,  # ← desde "Cant"
                cve_age=cve_age_val,
                nom_cte=nom_cte_val,
                rfc_cte=rfc_cte_val,
                des_mon=des_mon_val
            )
            
            # Evitar duplicados básicos por (fecha, codigo, cantidad)
            existe = Venta.query.filter_by(
                fecha=fecha,
                codigo=venta.codigo,
                cantidad=venta.cantidad
            ).first()
            if not existe:
                objetos.append(venta)
                nuevos += 1
        
        if objetos:
            db.session.bulk_save_objects(objetos)
            db.session.commit()
        
        flash(f"Archivo clasificado cargado. Nuevos: {nuevos}", "success")
        
    except Exception as e:
        db.session.rollback()
        app.logger.exception("[ventas_upload_clasificado] Error procesando archivo")
        flash(f"Error al procesar el clasificado: {e}", "error")
    
    return redirect(url_for("dashboard") + "#tab-ventas")

@app.route("/ventas/upload_hto", methods=["POST"])
@login_required
def ventas_upload_hto():
    archivo = request.files.get("archivo_excel")
    if not archivo:
        flash("No se subió ningún archivo HTO", "error")
        return redirect(url_for("dashboard") + "#tab-ventas")

    try:
        import re, math
        import numpy as np
        from unidecode import unidecode
        from sqlalchemy import or_

        # -------- Helpers anti-"nan" --------
        def clean_str(x) -> str:
            if x is None:
                return ""
            try:
                if isinstance(x, (float, np.floating)) and pd.isna(x):
                    return ""
            except Exception:
                pass
            s = str(x).strip()
            return "" if s.lower() in ("nan", "none", "null") else s

        def clean_float(x, default: float = 0.0) -> float:
            try:
                if x is None or (isinstance(x, (float, np.floating)) and pd.isna(x)):
                    return float(default)
                v = float(x)
                if math.isnan(v) or math.isinf(v):
                    return float(default)
                return v
            except Exception:
                return float(default)

        # --- Leer y normalizar encabezados
        df = pd.read_excel(archivo)

        def norm_hdr(s):
            s = unidecode(str(s)).strip().lower()
            s = re.sub(r"\(.*?\)", "", s)            # quita "(C3)" etc.
            s = re.sub(r"\s+", " ", s).strip()
            s = s.replace(" ", "_")
            return s

        df.columns = [norm_hdr(c) for c in df.columns]

        # --- Helper para elegir columnas tolerantes
        def pick(*candidates):
            cands = [norm_hdr(c) for c in candidates]
            for cand in cands:
                for col in df.columns:
                    if col == cand:
                        return col
            for cand in cands:
                for col in df.columns:
                    if cand in col:
                        return col
            return None

        # Campos posibles en HTO (acepta variantes)
        col_serie       = pick("serie")
        col_folio       = pick("folio", "no_fac", "numero", "num", "folio_factura")
        col_uuid        = pick("uuid_factura", "uuid", "uuid_cfdi")
        col_uuid_nc     = pick("uuid_nc", "uuid_relacion", "uuid_relacionado", "uuid_rel")
        col_cliente_c3  = pick("cliente_c3", "nombre_receptor", "cliente", "nombre")
        col_forma_pago  = pick("forma_de_pago_c3", "formadepago", "forma_de_pago")
        col_metodo_pago = pick("metodo_de_pago_c3", "metodo_de_pago")
        col_total_c3    = pick("total_c3", "total")
        col_pago        = pick("pago", "pago_1")

        if not (col_serie and col_folio):
            flash("HTO: no se detectaron las columnas de Serie/Folio.", "error")
            return redirect(url_for("dashboard") + "#tab-ventas")

        nuevos = 0

        for _, r in df.iterrows():
            serie = clean_str(r.get(col_serie))
            folio = clean_str(r.get(col_folio))
            if not serie or not folio:
                continue

            clave_codigo = (serie + folio).strip()

            # Buscar primero por codigo == Serie+Folio; si no, por no_fac == Folio
            venta = (
                Venta.query
                    .filter(or_(Venta.codigo == clave_codigo, Venta.no_fac == folio))
                    .order_by(Venta.id.desc())
                    .first()
            )
            if not venta:
                continue

            # Tomar valores del HTO (ya saneados)
            uuid_val   = clean_str(r.get(col_uuid)) if col_uuid else ""
            uuid_nc    = clean_str(r.get(col_uuid_nc)) if col_uuid_nc else ""
            cliente_c3 = clean_str(r.get(col_cliente_c3)) if col_cliente_c3 else ""
            forma_pago = clean_str(r.get(col_forma_pago)) if col_forma_pago else ""
            metodo     = clean_str(r.get(col_metodo_pago)) if col_metodo_pago else ""
            total_c3   = clean_float(_to_float(r.get(col_total_c3)) if col_total_c3 else 0.0)
            pago_1     = clean_str(r.get(col_pago)) if col_pago else ""

            # Autocompletar sin sobrescribir si ya hay dato
            if not venta.uuid_factura and uuid_val:
                venta.uuid_factura = uuid_val
            if not venta.uuid_nc and uuid_nc:
                venta.uuid_nc = uuid_nc
            if not venta.cliente_1 and cliente_c3:
                venta.cliente_1 = cliente_c3
            if not venta.forma_de_pago and forma_pago:
                venta.forma_de_pago = forma_pago
            if not venta.metodo_de_pago and metodo:
                venta.metodo_de_pago = metodo
            if (not venta.total_2 or venta.total_2 == 0) and total_c3:
                venta.total_2 = total_c3
            if not venta.pago_1 and pago_1:
                venta.pago_1 = pago_1

            nuevos += 1

        db.session.commit()
        flash(f"Archivo HTO cargado. Ventas actualizadas: {nuevos}", "success")

    except Exception as e:
        db.session.rollback()
        app.logger.exception("[ventas_upload_hto] Error procesando archivo HTO")
        flash(f"Error al procesar HTO: {e}", "error")

    return redirect(url_for("dashboard") + "#tab-ventas")





@app.route("/ventas/sumar_dia", methods=["GET"])
@login_required
def ventas_sumar_dia():
    fecha_str = (request.args.get("fecha") or "").strip()
    if not fecha_str:
        return {"ok": False, "msg": "Falta la fecha (YYYY-MM-DD)."}, 400
    
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except Exception:
        return {"ok": False, "msg": "Formato de fecha inválido (usa YYYY-MM-DD)."}, 400
    
    filtro_fecha = (Venta.fecha == fecha)
    # ✅ incluye nulos y excluye solo los que contengan devol
    filtro_no_devol = or_(Venta.subtipo.is_(None), ~Venta.subtipo.ilike("%DEVOL%"))
    
    # Más tolerante con efectivo
    tipo_norm = func.coalesce(Venta.tipo, '')
    subtipo_norm = func.coalesce(Venta.subtipo, '')
    filtro_efectivo = or_(tipo_norm.ilike("%EFECTIVO%"), subtipo_norm.ilike("%EFECTIVO%"))
    
    total_dia = db.session.query(func.coalesce(func.sum(Venta.cantidad), 0.0))\
        .filter(filtro_fecha, filtro_no_devol).scalar() or 0.0
    
    total_efectivo = db.session.query(func.coalesce(func.sum(Venta.cantidad), 0.0))\
        .filter(filtro_fecha, filtro_no_devol, filtro_efectivo).scalar() or 0.0
    
    # ✅ case con coalesce
    clasificacion = case(
        (subtipo_norm.ilike("%CAJA%5%MAYO%"), "CAJA 5 MAYO"),
        (subtipo_norm.ilike("%SLORENZ%"), "SAN LORENZO"),
        (subtipo_norm.ilike("%TEC%"), "TECNOLÓGICO"),
        (subtipo_norm.ilike("%PENIN%"), "PENINSULA"),
        else_="OTROS MOVIMIENTOS"
    )
    
    filas = db.session.query(
        clasificacion.label("clas"),
        func.coalesce(func.sum(Venta.cantidad), 0.0).label("monto")
    ).filter(
        filtro_fecha,
        filtro_no_devol
    ).group_by(
        clasificacion  # ← en vez de "clas"
    ).all()
    
    categorias = ["OTROS MOVIMIENTOS", "CAJA 5 MAYO", "SAN LORENZO", "TECNOLÓGICO", "PENINSULA"]
    desglose = {k: 0.0 for k in categorias}
    
    for clas, monto in filas:
        desglose[clas] = float(monto or 0.0)
    
    return {
        "ok": True,
        "fecha": fecha_str,
        "total_dia": float(total_dia),
        "total_efectivo": float(total_efectivo),
        "desglose": desglose
    }

@app.route("/ventas/eliminar_todo", methods=["POST"])
@login_required
def ventas_eliminar_todo():
    try:
        db.session.query(Venta).delete()
        db.session.commit()
        flash("Se eliminaron todas las ventas.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"No se pudieron eliminar las ventas: {e}", "error")
    
    return redirect(url_for("dashboard") + "#tab-ventas")

@app.route('/ventas')
@login_required
def filtrar_ventas():
    fecha_str = request.args.get('fecha_ventas')
    if fecha_str:
        return redirect(url_for('dashboard', fecha_ventas=fecha_str) + '#tab-ventas')
    return redirect(url_for('dashboard') + '#tab-ventas')

# Manejo de errores 500
@app.errorhandler(500)
def handle_500(error):
    db.session.rollback()  # Previene bloqueos de la base de datos
    return render_template('500.html'), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False') == 'True'
    app.run(host='0.0.0.0', port=port, debug=debug)